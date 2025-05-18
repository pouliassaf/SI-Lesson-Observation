#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 12 00:42:10 2025

@author: paulaassaf
"""

import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, timedelta, date, time # Import time specifically
import os
import statistics
import pandas as pd
import matplotlib.pyplot as plt
import csv
import io
from openpyxl.utils import get_column_letter

# Import ReportLab modules for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import re # Import regex for cleaning HTML tags
import numpy as np # Import numpy for isnan

# --- Set Streamlit Page Config (MUST BE THE FIRST STREAMLIT COMMAND) ---
st.set_page_config(page_title="Lesson Observation Tool", layout="wide")

# --- Logo File Paths ---
# Define a dictionary mapping school names to logo file paths
# Ensure these paths are correct relative to your script location
LOGO_PATHS = {
    "Default": "logos/default.jpeg", # Default logo for schools not listed
    "Al Bayan School": "logos/CS_Al Bayan Charter School_Logo.png",
    "Al Bayraq School": "logos/CS_Al Bayraq_Logo_PNG.png",
    "Al Dhaher School": "logos/CS_Al Dhaher_Logo_PNG.png",
    "Al Hosoon School": "logos/CS_Al Hosoon Charter School_Logo.png",
    "Al Mutanabi School": "logos/CS_Al Mutanabi_Logo_PNG.png",
    "Al Nahdha School": "logos/CS_Al Nahdha_Logo_PNG.png",
    "Jern Yafoor School": "logos/CS_Jern Yafoor_Logo_PNG.png",
    "Maryam Bint Omran School": "logos/CS_Maryam Bint Omran_Logo_PNG.png",
    # Add other school logos here
}

# --- Text Strings for Localization ---
en_strings = {
    "page_title": "Lesson Observation Tool",
    "sidebar_select_page": "Choose a page:",
    "page_lesson_input": "Lesson Observation Input",
    "page_analytics": "Lesson Observation Analytics",
    "page_help": "App Information and Guidelines",
    "title_lesson_input": "Weekly Lesson Observation Input Tool",
    "title_help": "App Information and Guidelines",
    "info_default_workbook": "Using default template workbook:",
    "warning_default_not_found": "Default template workbook '{}' not found. Please upload a workbook.",
    "error_opening_default": "Error opening default template file:",
    "success_lo_sheets_found": "Found {} LO sheets in workbook.",
    "select_sheet_or_create": "Select existing LO sheet or create a new one:",
    "option_create_new": "Create new",
    "success_sheet_created": "Created new sheet: {}",
    "error_template_not_found": "Error: 'LO 1' template sheet not found in the workbook! Cannot create new sheet.",
    "subheader_filling_data": "Filling data for: {}",
    "label_observer_name": "Observer Name",
    "label_teacher_name": "Teacher Name",
    "label_teacher_email": "Teacher Email",
    "label_operator": "Operator",
    "label_school_name": "School Name",
    "label_grade": "Grade",
    "label_subject": "Subject",
    "label_gender": "Gender",
    "label_students": "Number of Students",
    "label_males": "Number of Males",
    "label_females": "Number of Females",
    "label_time_in": "Time In",
    "label_time_out": "Time Out",
    "label_lesson_duration": "🕒 **Lesson Duration:** {} minutes — _{}_",
    "duration_full_lesson": "Full Lesson",
    "duration_walkthrough": "Walkthrough",
    "warning_calculate_duration": "Please enter both 'Time In' and 'Time Out' to calculate duration.",
    "warning_could_not_calculate_duration": "Could not calculate lesson duration.",
    "label_period": "Period",
    "label_obs_type": "Observation Type",
    "option_individual": "Individual",
    "option_joint": "Joint",
    "subheader_rubric_scores": "Rubric Scores",
    "expander_rubric_descriptors": "Rubric Guidance",
    "info_no_descriptors": "No rubric guidance available.",
    "label_rating_for": "Rating for {}",
    "label_write_notes": "Write notes for {}",
    "checkbox_send_feedback": "✉️ Generate Feedback Report (for PDF)",
    "button_save_observation": "💾 Save Observation",
    "warning_fill_essential": "Please fill in all essential information before saving.",
    "success_data_saved": "Observation data saved to workbook.",
    "error_saving_workbook": "Error saving workbook:",
    "download_workbook": "📥 Download updated workbook",
    "feedback_subject": "Lesson Observation Feedback",
    "feedback_greeting": "Dear {},\n\nYour lesson observation from {} has been saved.\n\n",
    "feedback_observer": "Observer: {}\n",
    "feedback_duration": "Duration: {}\n",
    "feedback_subject_fb": "Subject: {}\n",
    "feedback_school": "School: {}\n\n",
    "feedback_summary_header": "Here is a summary of your ratings based on the rubric:\n\n",
    "feedback_domain_header": "**{}: {}**\n",
    "feedback_element_rating": "- **{}:** Rating **{}**\n",
    "feedback_descriptor_for_rating": "  *Guidance for rating {}:* {}\n",
    "feedback_overall_score": "\n**Overall Average Score:** {:.2f}\n\n",
    "feedback_domain_average": "  *Domain Average:* {:.2f}\n",
    "feedback_performance_summary": "**Performance Summary:**\n",
    "overall_performance_level_text": "Overall Performance Level: {}",
    "feedback_domain_performance": "{}: {}\n",
    "feedback_support_plan_intro": "\n**Support Plan Recommended:**\n",
    "feedback_next_steps_intro": "\n**Suggested Next Steps:**\n",
    "feedback_closing": "\nBased on these ratings, please review your updated workbook for detailed feedback and areas for development.\n\n",
    "feedback_regards": "Regards,\nSchool Leadership Team",
    "success_feedback_generated": "Feedback generated.",
    "success_feedback_log_updated": "Feedback log updated.",
    "error_updating_log": "Error updating feedback log in workbook:",
    "title_analytics": "Lesson Observation Analytics Dashboard",
    "warning_no_lo_sheets_analytics": "No 'LO ' sheets found in the workbook for analytics.",
    "subheader_avg_score_overall": "Average Score per Domain (Across all observations)",
    "info_no_numeric_scores_overall": "No numeric scores found across all observations to calculate overall domain averages.",
    "subheader_data_summary": "Observation Data Summary",
    "subheader_filter_analyze": "Filter and Analyze",
    "filter_by_school": "Filter by School",
    "filter_by_grade": "Filter by Grade",
    "filter_by_subject": "Filter by Subject",
    "filter_by_operator": "Filter by Operator",
    "filter_by_observer_an": "Filter by Observer",
    "option_all": "All",
    "subheader_avg_score_filtered": "Average Score per Domain (Filtered)",
    "info_no_numeric_scores_filtered": "No observations matching the selected filters contain numeric scores for domain averages.",
    "subheader_observer_distribution": "Observer Distribution (Filtered)",
    "info_no_observer_data_filtered": "No observer data found for the selected filters.",
    "info_no_observation_data_filtered": "No observation data found for the selected filters.",
    "error_loading_analytics": "Error loading or processing workbook for analytics:",
    "overall_score_label": "Overall Score:",
    "overall_score_value": "**{:.2f}**",
    "overall_score_na": "**N/A**",
    "arabic_toggle_label": "عرض باللغة العربية (Display in Arabic)",
    "feedback_log_sheet_name": "Feedback Log", # Using English name for robustness with code logic
    "feedback_log_header": ["Sheet", "Observer", "Teacher", "Email", "School", "Subject", "Date", "Overall Judgment", "Overall Score", "Summary Notes"],
    "download_feedback_log_csv": "📥 Download Feedback Log (CSV)",
    "error_generating_log_csv": "Error generating Feedback Log CSV:",
    "download_overall_avg_csv": "📥 Download Overall Domain Averages (CSV)",
    "download_overall_avg_excel": "📥 Download Overall Domain Averages (Excel)",
    "download_filtered_avg_csv": "📥 Download Filtered Domain Averages (CSV)", # Not currently generated as a separate file
    "download_filtered_avg_excel": "📥 Download Filtered Domain Averages (Excel)", # Not currently generated as a separate file
    "download_filtered_data_csv": "📥 Download Filtered Visit Data (CSV)",
    "download_filtered_data_excel": "📥 Download Filtered Visit Data (Excel)",
    "label_observation_date": "Observation Date",
    "filter_start_date": "Start Date",
    "filter_end_date": "End Date",
    "filter_teacher": "Filter by Teacher",
    "subheader_teacher_performance": "Teacher Performance Over Time",
    "info_select_teacher": "Select a teacher to view individual performance analytics.",
    "info_no_obs_for_teacher": "No observations found for the selected teacher within the applied filters.",
    "subheader_teacher_domain_trend": "{} Domain Performance Trend",
    "subheader_teacher_overall_avg": "{} Overall Average Score (Filtered)",
    "perf_level_very_weak": "Very Weak",
    "perf_level_weak": "Weak",
    "perf_level_acceptable": "Acceptable",
    "perf_level_good": "Good",
    "perf_level_excellent": "Excellent",
    "plan_very_weak_overall": "Overall performance is Very Weak. A comprehensive support plan is required. Focus on fundamental teaching practices such as classroom management, lesson planning, and basic instructional strategies. Seek guidance from your mentor teacher and school leadership.",
    "plan_weak_overall": "Overall performance is Weak. A support plan is recommended. Identify 1-2 key areas for improvement from the observation and work with your mentor teacher to develop targeted strategies. Consider observing experienced colleagues in these areas.",
    "plan_weak_domain": "Performance in **{}** is Weak. Focus on developing skills related to: {}. Suggested actions include: [Specific action 1], [Specific action 2].", # Placeholder actions
    "steps_acceptable_overall": "Overall performance is Acceptable. Continue to build on your strengths. Identify one area for growth to refine your practice and enhance student learning.",
    "steps_good_overall": "Overall performance is Good. You demonstrate effective teaching practices. Explore opportunities to share your expertise with colleagues, perhaps through informal mentoring or presenting successful strategies.",
    "steps_good_domain": "Performance in **{}** is Good. You demonstrate strong skills in this area. Consider exploring advanced strategies related to: {}. Suggested actions include: [Specific advanced action 1], [Specific advanced action 2].", # Placeholder actions
    "steps_excellent_overall": "Overall performance is Excellent. You are a role model for effective teaching. Consider leading professional development sessions or mentoring less experienced teachers.",
    "steps_excellent_domain": "Performance in **{}** is Excellent. Your practice in this area is exemplary. Continue to innovate and refine your practice, perhaps by researching and implementing cutting-edge strategies related to: {}.", # Placeholder actions
    "no_specific_plan_needed": "Performance is at an acceptable level or above. No immediate support plan required based on this observation. Focus on continuous improvement based on your professional goals.",
    "warning_fill_basic_info": "Please fill in Observer Name, Teacher Name, School Name, Grade, Subject, Gender, and Observation Date.",
    "warning_fill_all_basic_info": "Please fill in all basic information fields.", # Generic fallback - removed specific fields for simplicity
    "warning_numeric_fields": "Please enter valid numbers for Students, Males, and Females.",
    "success_pdf_generated": "Feedback PDF generated successfully.",
    "download_feedback_pdf": "📥 Download Feedback PDF",
    "checkbox_cleanup_sheets": "🪟 Clean up unused LO sheets (no observer name)",
    "warning_sheets_removed": "Removed {} unused LO sheets.",
    "info_reloaded_workbook": "Reloaded workbook after cleanup.",
    "info_no_sheets_to_cleanup": "No unused LO sheets found to clean up.",
    "expander_guidelines": "📘 Click here to view observation guidelines",
    "info_no_guidelines": "Guidelines sheet is empty or could not be read.",
    "warning_select_create_sheet": "Please select or create a valid sheet to proceed.",
    "label_overall_notes": "General Notes for this Lesson Observation",

}

# Placeholder Arabic strings - REPLACE THESE WITH ACTUAL TRANSLATIONS
# NOTE: Arabic support in ReportLab PDFs requires Arabic fonts and potentially bi-directional text handling.
# The translations below are placeholders and ReportLab PDF output for these may not render correctly without additional setup.
ar_strings = {
    "page_title": "أداة تقييم الزيارات الصفية",
    "sidebar_select_page": "اختر صفحة:",
    "page_lesson_input": "إدخال تقييم الزيارة",
    "page_analytics": "تحليلات الزيارات الصفية",
    "page_help": "معلومات وإرشادات التطبيق",
    "title_lesson_input": "أداة إدخال زيارة صفية أسبوعية",
    "title_help": "معلومات وإرشادات التطبيق",
    "info_default_workbook": "استخدام مصنف القالب الافتراضي:",
    "warning_default_not_found": "تحذير: لم يتم العثور على مصنف القالب الافتراضي '{}'. يرجى تحميل مصنف.",
    "error_opening_default": "خطأ في فتح ملف القالب الافتراضي:",
    "success_lo_sheets_found": "تم العثور على {} أوراق LO في المصنف.",
    "select_sheet_or_create": "حدد ورقة LO موجودة أو أنشئ واحدة جديدة:",
    "option_create_new": "إنشاء جديد",
    "success_sheet_created": "تم إنشاء ورقة جديدة: {}",
    "error_template_not_found": "خطأ: لم يتم العثور على ورقة القالب 'LO 1' في المصنف! لا يمكن إنشاء ورقة جديدة.",
    "subheader_filling_data": "ملء البيانات لـ: {}",
    "label_observer_name": "اسم المراقب",
    "label_teacher_name": "اسم المعلم",
    "label_teacher_email": "البريد الإلكتروني للمعلم",
    "label_operator": "المشغل",
    "label_school_name": "اسم المدرسة",
    "label_grade": "الصف",
    "label_subject": "المادة",
    "label_gender": "الجنس",
    "label_students": "عدد الطلاب",
    "label_males": "عدد الذكور",
    "label_females": "عدد الإناث",
    "label_time_in": "وقت الدخول",
    "label_time_out": "وقت الخروج",
    "label_lesson_duration": "🕒 **مدة الدرس:** {} دقيقة — _{}_",
    "duration_full_lesson": "درس كامل",
    "duration_walkthrough": "جولة سريعة",
    "warning_calculate_duration": "يرجى إدخال وقت الدخول ووقت الخروج لحساب المدة.",
    "warning_could_not_calculate_duration": "تعذر حساب مدة الدرس.",
    "label_period": "الفترة",
    "label_obs_type": "نوع الزيارة",
    "option_individual": "فردي",
    "option_joint": "مشترك",
    "subheader_rubric_scores": "درجات الدليل",
    "expander_rubric_descriptors": "إرشادات الدليل",
    "info_no_descriptors": "لا توجد إرشادات دليل متاحة لهذا العنصر.",
    "label_rating_for": "التقييم لـ {}",
    "label_write_notes": "كتابة ملاحظات لـ {}",
    "checkbox_send_feedback": "✉️ إنشاء تقرير الملاحظات (للملف PDF)",
    "button_save_observation": "💾 حفظ الزيارة",
    "warning_fill_essential": "يرجى ملء جميع حقول المعلومات الأساسية قبل الحفظ.",
    "success_data_saved": "تم حفظ بيانات الزيارة في المصنف.",
    "error_saving_workbook": "خطأ في حفظ المصنف:",
    "download_workbook": "📥 تنزيل المصنف المحدث",
    "feedback_subject": "ملاحظات الزيارة الصفية",
    "feedback_greeting": "عزيزي {},\n\nتم حفظ زيارتك الصفية من {}.\n\n",
    "feedback_observer": "المراقب: {}\n",
    "feedback_duration": "المدة: {}\n",
    "feedback_subject_fb": "المادة: {}\n",
    "feedback_school": "المدرسة: {}\n\n",
    "feedback_summary_header": "إليك ملخص لتقييماتك بناءً على الدليل:\n\n",
    "feedback_domain_header": "**{}: {}**\n",
    "feedback_element_rating": "- **{}:** التقييم **{}**\n",
    "feedback_descriptor_for_rating": "  *إرشادات للتقييم {}:* {}\n",
    "feedback_overall_score": "\n**متوسط الدرجة الإجمالي:** {:.2f}\n\n",
    "feedback_domain_average": "  *متوسط المجال:* {:.2f}\n",
    "feedback_performance_summary": "**ملخص الأداء:**\n",
    "overall_performance_level_text": "مستوى الأداء الإجمالي: {}",
    "feedback_domain_performance": "{}: {}\n",
    "feedback_support_plan_intro": "\n**خطة الدعم الموصى بها:**\n",
    "feedback_next_steps_intro": "\n**الخطوات التالية المقترحة:**\n",
    "feedback_closing": "\nبناءً على هذه التقييمات، يرجى مراجعة المصنف المحدث للحصول على ملاحظات تفصيلية ومجالات التطوير.\n\n",
    "feedback_regards": "مع التحيات,\nفريق قيادة المدرسة",
    "success_feedback_generated": "تم إنشاء الملاحظات.",
    "success_feedback_log_updated": "تم تحديث سجل الملاحظات.",
    "error_updating_log": "خطأ في تحديث سجل الملاحظات في المصنف:",
    "title_analytics": "لوحة تحليلات الزيارة الصفية",
    "warning_no_lo_sheets_analytics": "لم يتم العثور على أوراق 'LO ' في المصنف للتحليلات.",
    "subheader_avg_score_overall": "متوسط الدرجة لكل مجال (عبر جميع الزيارات)",
    "info_no_numeric_scores_overall": "لم يتم العثور على درجات رقمية عبر جميع الزيارات لحساب متوسطات المجال الإجمالية.",
    "subheader_data_summary": "ملخص بيانات الزيارة",
    "subheader_filter_analyze": "تصفية وتحليل",
    "filter_by_school": "تصفية حسب المدرسة",
    "filter_by_grade": "تصفية حسب الصف",
    "filter_by_subject": "تصفية حسب المادة",
    "filter_by_operator": "تصفية حسب المشغل",
    "filter_by_observer_an": "تصفية حسب المراقب",
    "option_all": "الكل",
    "subheader_avg_score_filtered": "متوسط الدرجة لكل مجال (مصفى)",
    "info_no_numeric_scores_filtered": "لا توجد زيارات مطابقة للمرشحات المحددة تحتوي على درجات رقمية لمتوسطات المجال.",
    "subheader_observer_distribution": "توزيع المراقبين (مصفى)",
    "info_no_observer_data_filtered": "لم يتم العثور على بيانات المراقب للمرشحات المحددة.",
    "info_no_observation_data_filtered": "لم يتم العثور على بيانات الزيارة للمرشحات المحددة.",
    "error_loading_analytics": "خطأ في تحميل أو معالجة المصنف للتحليلات:",
    "overall_score_label": "النتيجة الإجمالية:",
    "overall_score_value": "**{:.2f}**",
    "overall_score_na": "**غير متوفر**",
    "arabic_toggle_label": "عرض باللغة العربية (Display in Arabic)",
    "feedback_log_sheet_name": "سجل الملاحظات", # Using Arabic name for robustness
    "feedback_log_header": ["Sheet", "Observer", "Teacher", "Email", "School", "Subject", "Date", "Overall Judgment", "Overall Score", "Summary Notes"], # Keep English keys for code, display Arabic in UI if needed
    "download_feedback_log_csv": "📥 تنزيل سجل الملاحظات (CSV)",
    "error_generating_log_csv": "خطأ في إنشاء سجل الملاحظات CSV:",
    "download_overall_avg_csv": "📥 تنزيل متوسطات المجال الإجمالية (CSV)",
    "download_overall_avg_excel": "📥 تنزيل متوسطات المجال الإجمالية (Excel)",
    "download_filtered_avg_csv": "📥 تنزيل متوسطات المجال المصفاة (CSV)", # Not currently generated as a separate file
    "download_filtered_avg_excel": "📥 تنزيل متوسطات المجال المصفاة (Excel)", # Not currently generated as a separate file
    "download_filtered_data_csv": "📥 تنزيل بيانات الزيارة المصفاة (CSV)",
    "download_filtered_data_excel": "📥 تنزيل بيانات الزيارة المصفاة (Excel)",
    "label_observation_date": "تاريخ الزيارة",
    "filter_start_date": "تاريخ البدء",
    "filter_end_date": "تاريخ الانتهاء",
    "filter_teacher": "تصفية حسب المعلم",
    "subheader_teacher_performance": "أداء المعلم بمرور الوقت",
    "info_select_teacher": "حدد معلمًا لعرض تحليلات الأداء الفردي.",
    "info_no_obs_for_teacher": "لم يتم العثور على زيارات للمعلم المحدد ضمن المرشحات المطبقة.",
    "subheader_teacher_domain_trend": "اتجاه أداء مجال {}",
    "subheader_teacher_overall_avg": "متوسط الدرجة الإجمالي لـ {} (مصفى)",
    "perf_level_very_weak": "ضعيف جداً",
    "perf_level_weak": "ضعيف",
    "perf_level_acceptable": "مقبول",
    "perf_level_good": "جيد",
    "perf_level_excellent": "ممتاز",
    "plan_very_weak_overall": "الأداء الإجمالي ضعيف جداً. تتطلب خطة دعم شاملة. ركز على الممارسات التعليمية الأساسية مثل إدارة الصف، وتخطيط الدرس، والاستراتيجيات التعليمية الأساسية. اطلب التوجيه من معلمك الموجه وقيادة المدرسة.",
    "plan_weak_overall": "الأداء الإجمالي ضعيف. يوصى بخطة دعم. حدد 1-2 من المجالات الرئيسية للتحسين من الملاحظة واعمل مع معلمك الموجه لتطوير استراتيجيات مستهدفة. فكر في ملاحظة الزملاء ذوي الخبرة في هذه المجالات.",
    "plan_weak_domain": "الأداء في **{}** ضعيف. ركز على تطوير المهارات المتعلقة بـ: {}. الإجراءات المقترحة تشمل: [إجراء محدد 1]، [إجراء محدد 2].",
    "steps_acceptable_overall": "الأداء الإجمالي مقبول. استمر في البناء على نقاط قوتك. حدد مجالًا واحدًا للنمو لتحسين ممارستك وتعزيز تعلم الطلاب.",
    "steps_good_overall": "الأداء الإجمالي جيد. أنت تظهر ممارسات تعليمية فعالة. استكشف فرص مشاركة خبرتك مع الزملاء، ربما من خلال التوجيه غير الرسمي أو تقديم استراتيجيات ناجحة.",
    "steps_good_domain": "الأداء في **{}** جيد. أنت تظهر مهارات قوية في هذا المجال. فكر في استكشاف استراتيجيات متقدمة تتعلق بـ: {}. الإجراءات المقترحة تشمل: [إجراء متقدم محدد 1]، [إجراء متقدم محدد 2].",
    "steps_excellent_overall": "الأداء الإجمالي ممتاز. أنت نموذج يحتذى به في التدريس الفعال. فكر في قيادة جلسات التطوير المهني أو توجيه المعلمين الأقل خبرة.",
    "steps_excellent_domain": "الأداء في **{}** ممتاز. ممارستك في هذا المجال نموذجية. استمر في الابتكار وتحسين ممارستك، ربما من خلال البحث وتطبيق استراتيجيات حديثة تتعلق بـ: {}.",
    "no_specific_plan_needed": "الأداء عند مستوى مقبول أو أعلى. لا توجد خطة دعم فورية مطلوبة بناءً على هذه الملاحظة. ركز على التحسين المستمر بناءً على أهدافك المهنية.",
    "warning_fill_essential": "يرجى ملء جميع حقول المعلومات الأساسية قبل الحفظ.", # Simplified generic warning
    "warning_numeric_fields": "يرجى إدخال أرقام صحيحة لحقول الطلاب، الذكور، والإناث.",
    "success_pdf_generated": "تم إنشاء ملف الملاحظات PDF بنجاح.",
    "download_feedback_pdf": "📥 تنزيل ملف الملاحظات PDF",
    "checkbox_cleanup_sheets": "🪟 تنظيف أوراق LO غير المستخدمة (لا يوجد اسم مراقب)",
    "warning_sheets_removed": "تمت إزالة {} أوراق LO غير مستخدمة.",
    "info_reloaded_workbook": "تمت إعادة تحميل المصنف بعد التنظيف.",
    "info_no_sheets_to_cleanup": "لم يتم العثور على أوراق LO غير مستخدمة لتنظيفها.",
    "expander_guidelines": "📘 انقر هنا لعرض إرشادات الملاحظة",
    "info_no_guidelines": "ورقة الإرشادات فارغة أو تعذر قراءتها.",
    "warning_select_create_sheet": "يرجى تحديد أو إنشاء ورقة صالحة للمتابعة.",
    "label_overall_notes": "ملاحظات عامة لهذه الملاحظة الصفية",
}


# --- Function to get strings based on language toggle ---
def get_strings(arabic_mode):
    return ar_strings if arabic_mode else en_strings

# --- Function to determine performance level based on score ---
def get_performance_level(score, strings):
    # Ensure score is treated as a number for comparison, handle None/NaN/non-numeric
    if score is None or (isinstance(score, (int, float)) and math.isnan(score)) or not isinstance(score, (int, float)):
         return strings["overall_score_na"]

    try:
        numeric_score = float(score)
        if numeric_score >= 5.5:
            return strings["perf_level_excellent"]
        elif numeric_score >= 4.5:
            return strings["perf_level_good"]
        elif numeric_score >= 3.5:
            return strings["perf_level_acceptable"]
        elif numeric_score >= 2.5:
            return strings["perf_level_weak"]
        else:
            return strings["perf_level_very_weak"]
    except (ValueError, TypeError):
        # This catch should ideally not be hit if the initial check works, but included for safety
        return strings["overall_score_na"]


# --- Define ReportLab Styles (DEFINED ONCE) ---
# Get default stylesheet
styles = getSampleStyleSheet()

# Add custom styles if they don't exist
if 'Heading1Centered' not in styles:
    styles.add(ParagraphStyle(name='Heading1Centered', alignment=1, fontSize=16, spaceAfter=14, bold=1))
if 'Heading2' not in styles:
    styles.add(ParagraphStyle(name='Heading2', fontSize=12, spaceAfter=10, bold=1))
if 'Normal' not in styles:
    styles.add(ParagraphStyle(name='Normal', fontSize=10, spaceAfter=6))
if 'RubricDescriptor' not in styles:
    styles.add(ParagraphStyle(name='RubricDescriptor', fontSize=9, spaceAfter=4, leftIndent=18)) # Indent descriptors
if 'RubricDomainHeading' not in styles:
    styles.add(ParagraphStyle(name='RubricDomainHeading', fontSize=11, spaceAfter=8, bold=1)) # Style for domain headings in PDF
if 'RubricElementRating' not in styles:
    styles.add(ParagraphStyle(name='RubricElementRating', fontSize=10, spaceAfter=4, leftIndent=10)) # Style for element rating in PDF

# --- Function to generate PDF ---
def generate_observation_pdf(data, feedback_content, strings): # Removed teacher_email as it's in data now
    buffer = io.BytesIO()
    # Note: For Arabic support in PDF, you need to register Arabic fonts with ReportLab
    # and potentially adjust directionality in styles. This requires more advanced ReportLab setup.
    # The current setup will likely render Arabic characters incorrectly or as boxes.
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    story = []

    # --- Add School Logo ---
    school_name = data.get("school_name", "Default") # Use key from data dict
    logo_path = LOGO_PATHS.get(school_name, LOGO_PATHS["Default"])

    # Add logo only if file exists and is an image
    is_image_logo = False
    if os.path.exists(logo_path):
         try:
             # Basic check for image extension, not foolproof
             if logo_path.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                 img = Image(logo_path, width=1.5*inch, height=0.75*inch)
                 img.hAlign = 'CENTER'
                 story.append(img)
                 story.append(Spacer(1, 0.2*inch))
                 is_image_logo = True
             else:
                 # Not a recognized image format, add text title instead
                 story.append(Paragraph(strings["page_title"], styles['Heading1Centered']))
                 story.append(Spacer(1, 0.1*inch))
                 print(f"Logo file {logo_path} is not a standard image format. Using text title.")
         except Exception as e:
              print(f"Could not add logo for {school_name} from {logo_path}: {e}")
              story.append(Paragraph(f"[{school_name} Logo Placeholder]", styles['Normal'])) # Add placeholder text
              story.append(Spacer(1, 0.1*inch))
              # Fallback to text title if logo failed
              story.append(Paragraph(strings["page_title"], styles['Heading1Centered']))
              story.append(Spacer(1, 0.1*inch))
    else:
        print(f"Logo file not found for {school_name} at {logo_path}. Using text title.")
        story.append(Paragraph(strings["page_title"], styles['Heading1Centered']))
        story.append(Spacer(1, 0.1*inch))

    # Basic Information Table - Update keys to match data dictionary from input fields
    # Collect all basic info first to decide what to include
    basic_info_map = {
        strings["label_observer_name"]: data.get("observer_name", ""),
        strings["label_teacher_name"]: data.get("teacher_name", ""),
        strings["label_teacher_email"]: data.get("teacher_email", ""),
        strings["label_operator"]: data.get("operator", ""),
        strings["label_school_name"]: data.get("school_name", ""),
        strings["label_grade"]: data.get("grade", ""),
        strings["label_subject"]: data.get("subject", ""),
        strings["label_gender"]: data.get("gender", ""),
        strings["label_students"]: data.get("students", ""),
        strings["label_males"]: data.get("males", ""),
        strings["label_females"]: data.get("females", ""),
        strings["label_observation_date"]: data.get("observation_date", ""),
        strings["label_time_in"]: data.get("time_in", ""),
        strings["label_time_out"]: data.get("time_out", ""),
        # The duration label needs care with HTML/emoji
        strings["label_lesson_duration"].split("🕒")[0].strip(): data.get("duration_display", ""), # Pass formatted duration, strip emoji/html
        strings["label_period"]: data.get("period", ""),
        strings["label_obs_type"]: data.get("observation_type", ""),
    }

    # Format basic info data for the table, excluding None/empty values
    cleaned_basic_info_data = []
    # Ensure the Overall Score label and value are always last if score is available
    overall_score_pdf_row = None
    if data.get("overall_score_display") and data.get("overall_score_display") != strings["overall_score_na"]:
         overall_score_pdf_row = [strings["overall_score_label"] + ":", data.get("overall_score_display", strings["overall_score_na"])]


    for label, value in basic_info_map.items():
        # Only include if value is not None, not an empty string, and not NaN
        if value is not None and (not isinstance(value, str) or value.strip() != "") and not (isinstance(value, float) and math.isnan(value)):
             formatted_value = value
             # Format date/time objects nicely
             if isinstance(value, datetime):
                 formatted_value = value.strftime("%Y-%m-%d %H:%M")
             elif isinstance(value, date):
                 formatted_value = value.strftime("%Y-%m-%d")
             elif isinstance(value, time): # Use time object
                  formatted_value = value.strftime("%H:%M")
             elif isinstance(value, (int, float)):
                  formatted_value = str(value) # Convert numbers to string


             cleaned_basic_info_data.append([str(label) + ":", str(formatted_value)]) # Ensure keys are strings

    # Add the overall score row at the end if it exists
    if overall_score_pdf_row:
         cleaned_basic_info_data.append(overall_score_pdf_row)


    if cleaned_basic_info_data: # Only add table if there is data to display
         table_style = TableStyle([
             ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
             ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
             ('FONTSIZE', (0, 0), (-1, 0), 10),
             ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
             ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
             ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
             ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
             ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
             ('FONTSIZE', (0, 1), (-1, -1), 9),
             ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
             ('GRID', (0, 0), (-1, -1), 1, colors.black),
             ('BOX', (0, 0), (-1, -1), 1, colors.black),
         ])

         # Calculate optimal column widths: 30% for label, 70% for value
         available_width = letter[0] - (doc.leftMargin + doc.rightMargin)
         col_widths = [available_width * 0.3, available_width * 0.7]


         table = Table(cleaned_basic_info_data, colWidths=col_widths)
         table.setStyle(table_style)
         story.append(table)
         story.append(Spacer(1, 0.2*inch))


    # Rubric Scores - Needs to be adapted to the data structure collected from the inputs
    story.append(Paragraph(strings["subheader_rubric_scores"], styles['Heading2']))

    # Assuming 'data' dictionary passed to the PDF function contains domain_data
    domain_data = data.get("domain_data", {})

    if domain_data:
        for domain_name, domain_info in domain_data.items():
            domain_title = domain_info.get("title", domain_name)
            domain_average = domain_info.get("average")
            domain_judgment = domain_info.get("judgment")
            elements = domain_info.get("elements", [])

            # Domain Title and Average
            story.append(Paragraph(f"<b>{domain_name}: {domain_title}</b>", styles['RubricDomainHeading']))
            if domain_average is not None and not math.isnan(domain_average):
                 story.append(Paragraph(f"  Domain Average: {domain_average:.2f} ({domain_judgment})", styles['Normal']))
            else:
                 story.append(Paragraph(f"  Domain Average: {strings['overall_score_na']}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))

            # Elements
            if elements:
                 for element in elements:
                     label = element.get("label", "Unknown Element")
                     rating = element.get("rating", "N/A")
                     note = element.get("note", "")
                     descriptor = element.get("descriptor", "") # Specific descriptor text for the given rating

                     story.append(Paragraph(f"- <b>{label}:</b> Rating <b>{rating}</b>", styles['RubricElementRating']))

                     # Include Descriptor if available
                     if descriptor and isinstance(descriptor, str) and descriptor.strip():
                         # Clean and format the descriptor text - remove HTML/markdown like bold
                         cleaned_desc_para = re.sub(r'<.*?>', '', descriptor).replace('**', '')
                         desc_paragraphs = cleaned_desc_para.split('\n')
                         for desc_para in desc_paragraphs:
                             if desc_para.strip():
                                  story.append(Paragraph(desc_para, styles['RubricDescriptor']))
                         story.append(Spacer(1, 0.05*inch)) # Small space after descriptor

                     # Include Note if available
                     if note and isinstance(note, str) and note.strip():
                          # Clean and format the note text - remove HTML/markdown like bold
                          cleaned_note_para = re.sub(r'<.*?>', '', note).replace('**', '')
                          note_paragraphs = cleaned_note_para.split('\n')
                          story.append(Paragraph("  <i>Notes:</i>", styles['Normal'])) # Italicize notes header
                          for note_para in note_paragraphs:
                             if note_para.strip():
                                  story.append(Paragraph(note_para, styles['Normal']))
                          story.append(Spacer(1, 0.05*inch)) # Small space after notes

                     story.append(Spacer(1, 0.1*inch)) # Space after each element
            else:
                 story.append(Paragraph("No elements recorded for this domain.", styles['Normal']))


            story.append(Spacer(1, 0.2*inch)) # Space after each domain
    else:
        story.append(Paragraph("No rubric data available.", styles['Normal']))


    # Add Overall Notes
    overall_notes = data.get("overall_notes", "")
    if overall_notes and isinstance(overall_notes, str) and overall_notes.strip():
        story.append(Paragraph(strings["label_overall_notes"] + ":", styles['Heading2']))
        # Convert markdown in general notes if any
        cleaned_overall_notes = re.sub(r'<.*?>', '', overall_notes).replace('**', '<b>').replace('**', '</b>')
        notes_paragraphs = cleaned_overall_notes.split('\n')
        for note_para in notes_paragraphs:
            if note_para.strip():
                 story.append(Paragraph(note_para, styles['Normal']))
        story.append(Spacer(1, 0.2*inch))


    # Feedback Content (This part is crucial and needs to be generated dynamically)
    story.append(Paragraph("Feedback Report:", styles['Heading2']))
    if feedback_content and isinstance(feedback_content, str):
        # The feedback_content string needs to be constructed before calling this function.
        # It should include the greeting, summary of scores/judgments, performance summary,
        # suggested plan/steps, and closing.
        # Convert markdown-like text to ReportLab flowables
        feedback_paragraphs = feedback_content.split('\n\n') # Split by double newline
        for para in feedback_paragraphs:
            if para.strip():
                # Simple bold conversion and newline handling for ReportLab
                # Also clean any stray HTML tags
                para_styled = re.sub(r'<.*?>', '', para).replace('**', '<b>').replace('**', '</b>').replace('\n', '<br/>')
                story.append(Paragraph(para_styled, styles['Normal']))
            story.append(Spacer(1, 0.1*inch)) # Add space between paragraphs
    else:
        story.append(Paragraph("Feedback content could not be generated.", styles['Normal']))


    # Build the PDF
    try:
        doc.build(story)
        buffer.seek(0)
        return buffer
    except Exception as e:
        # Handle errors during PDF build
        print(f"Error generating PDF: {e}") # Log to console/logs
        # st.error(f"Error generating PDF: {e}") # Avoid st.error inside this function if it runs outside Streamlit thread
        return None # Indicate failure


# --- Streamlit App Layout ---
# Add Arabic toggle early to affect language throughout the app
arabic_mode = st.sidebar.toggle(en_strings["arabic_toggle_label"], False)
strings = get_strings(arabic_mode)

# Initialize workbook in session state if not already loaded
# This avoids reloading the workbook from disk on every single rerun
# BUT requires careful handling of saving modifications back to disk.
if 'workbook' not in st.session_state:
    st.session_state.workbook = None
    # Attempt to load the default workbook on first run
    if os.path.exists(DEFAULT_FILE):
        try:
            st.session_state.workbook = load_workbook(DEFAULT_FILE)
            st.success(strings["info_default_workbook"].format(DEFAULT_FILE))
        except Exception as e:
            st.error(strings["error_opening_default"].format(e))
            st.session_state.workbook = None
    else:
        st.warning(strings["warning_default_not_found"].format(DEFAULT_FILE))


# Use the workbook from session state
wb = st.session_state.workbook

# Sidebar page selection
page = st.sidebar.selectbox(strings["sidebar_select_page"], [strings["page_lesson_input"], strings["page_analytics"], strings["page_help"]]) # Added Help page

# --- Main Application Logic ---
if wb: # Proceed only if workbook was loaded successfully
    if page == strings["page_lesson_input"]:
        st.title(strings["title_lesson_input"])

        # CSS for spacing
        st.markdown("""
        <style>
        .block-container {
            padding-top: 2rem;
        }
        </style>
        """, unsafe_allow_html=True)

        # Email Domain Restriction
        # Only show the rest of the app if email is entered and authorized
        email = st.text_input("Enter your school email to continue", value=st.session_state.get('auth_email_input', ''), key='auth_email_input')
        allowed_domains = ["@charterschools.ae", "@adek.gov.ae"]
        # Check if email is entered AND if it ends with an allowed domain
        if not (email and any(email.strip().lower().endswith(domain) for domain in allowed_domains)):
             if email.strip(): # Only show specific warning if email is entered but invalid
                  st.warning("Access restricted. Please use an authorized school email.")
             st.stop() # This stops the rest of the script below this point from running

        # If email is valid, ensure it's stored and show the rest of the content
        st.session_state['auth_email_input'] = email # Store valid email


        lo_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("LO ")]
        # Only report LO sheets if workbook loaded successfully and there are LO sheets
        if wb and lo_sheets:
             st.success(strings["success_lo_sheets_found"].format(len(lo_sheets)))

        # Cleanup unused LO sheets
        # Only show cleanup option if there's more than just the template sheet
        # and workbook is loaded
        if wb and len(lo_sheets) > 1:
             if st.checkbox(strings.get("checkbox_cleanup_sheets", "🪟 Clean up unused LO sheets (no observer name)")):
                 to_remove = []
                 # Use AA1 as indicator for data existence (Observer Name)
                 for sheet in lo_sheets:
                     # Don't attempt to clean up the template sheet
                     if sheet == "LO 1":
                         continue
                     try:
                         # Check AA1 value in the sheet - if None or empty string, mark for removal
                         aa1_value = wb[sheet]["AA1"].value
                         if aa1_value is None or (isinstance(aa1_value, str) and aa1_value.strip() == ""):
                             to_remove.append(sheet)
                     except KeyError:
                         # If AA1 doesn't exist, consider it unused/corrupt
                         to_remove.append(sheet)
                     except Exception as e:
                         st.warning(f"Could not check sheet '{sheet}' for cleanup: {e}")


                 if to_remove:
                     st.info(strings.get("warning_sheets_removed", "Removed {} unused LO sheets.").format(len(to_remove)))
                     for sheet in to_remove:
                         # Double check it's not the template and still exists before removing
                         if sheet != "LO 1" and sheet in wb.sheetnames:
                             try:
                                 wb.remove(wb[sheet])
                             except Exception as e:
                                 st.error(f"Error removing sheet {sheet}: {e}") # Report removal errors

                     # Save the workbook immediately after cleanup
                     try:
                          wb.save(DEFAULT_FILE)
                          st.success("Workbook saved after cleanup.")
                          # Reload the workbook into session state after saving changes
                          st.session_state.workbook = load_workbook(DEFAULT_FILE)
                          wb = st.session_state.workbook # Update the local wb variable
                          st.info(strings.get("info_reloaded_workbook", "Reloaded workbook after cleanup."))
                          # Re-run Streamlit to update the sheet selector dropdown
                          st.rerun()
                     except Exception as e:
                          st.error(strings["error_saving_workbook"].format(e))

                 else:
                     st.info(strings.get("info_no_sheets_to_cleanup", "No unused LO sheets found to clean up."))


        # Display Guidelines
        # Moved the expander to the Help page primarily, but can optionally keep it here
        # Keeping it here allows quick access on the input page too.
        if wb and "Guidelines" in wb.sheetnames: # Ensure workbook is loaded before checking sheet
            guideline_content = []
            try:
                # Read cells row by row, value only, skip None
                for row in wb["Guidelines"].iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            # Convert potential numbers to string and strip whitespace
                            guideline_content.append(str(cell).strip())
            except Exception as e:
                st.error(f"Error reading Guidelines sheet: {e}")
                guideline_content = [f"Error loading guidelines: {e}"] # Provide an error message

            # Join only non-empty lines
            cleaned_guidelines = [line for line in guideline_content if line]
            if cleaned_guidelines:
                st.expander(strings.get("expander_guidelines", "📘 Click here to view observation guidelines")).markdown(
                    "\n".join(cleaned_guidelines) # Join with newline for markdown
                )
            else:
                st.info(strings.get("info_no_guidelines", "Guidelines sheet is empty or could not be read."))


        lo_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("LO ")]
        # Ensure "LO 1" template is always available for copying
        if "LO 1" not in wb.sheetnames:
            st.error(strings["error_template_not_found"])
            st.stop() # Cannot proceed without template

        # Add "Create new" option only if "LO 1" exists and workbook is loaded
        # The LO 1 sheet should generally not be selectable for input, only used as a template.
        # So we only list existing LO sheets (excluding LO 1) and the "Create new" option.
        existing_sheets_for_selection = sorted([s for s in lo_sheets if s != "LO 1"])
        sheet_selection_options = [strings["option_create_new"]] + existing_sheets_for_selection

        # Determine initial selection index (try to keep current sheet if exists)
        # Reset to 'Create new' if the previously selected sheet was just cleaned up or is template
        # Use a unique key for the selectbox to ensure its state is managed correctly
        selected_option = st.selectbox(
            strings["select_sheet_or_create"],
            sheet_selection_options,
            key='sheet_selector' # Use a session state key
        )


        sheet_name_to_process = None
        ws_to_load_from = None # Initialize worksheet to load data from

        # --- Function to read existing data from a sheet (to pre-fill inputs) ---
        def load_existing_data(worksheet: Worksheet, rubric_structure):
            data = {}
            # Basic Info from snippet 2 save locations
            try: data["grade"] = worksheet["B1"].value # Assuming Grade is in B1
            except Exception: pass
            try: data["gender"] = worksheet["B5"].value
            except Exception: pass
            try: data["students"] = worksheet["B6"].value
            except Exception: pass
            try: data["males"] = worksheet["B7"].value
            except Exception: pass
            try: data["females"] = worksheet["B8"].value
            except Exception: pass
            try: data["subject"] = worksheet["D2"].value
            except Exception: pass

            # Date from assumed location D10
            try:
                date_val = worksheet["D10"].value
                if isinstance(date_val, datetime):
                    data["observation_date"] = date_val.date() # Store as date object
                elif isinstance(date_val, date):
                    data["observation_date"] = date_val # Already a date object
                # Handle potential string dates if needed:
                elif isinstance(date_val, str) and date_val:
                     try:
                          data["observation_date"] = datetime.strptime(date_val, "%Y-%m-%d").date()
                     except ValueError:
                          pass # Keep default or None if parsing fails
            except Exception:
                pass # Default date will be handled by the input widget

            # Duration was calculated, need time in/out
            try:
                time_in_val = worksheet["D7"].value
                if isinstance(time_in_val, datetime.time):
                    data["time_in"] = time_in_val
                elif isinstance(time_in_val, datetime): # openpyxl sometimes reads time as datetime
                    data["time_in"] = time_in_val.time()
                elif isinstance(time_in_val, str) and time_in_val:
                    try:
                        data["time_in"] = datetime.strptime(time_in_val, "%H:%M:%S").time()
                    except ValueError:
                         try:
                              data["time_in"] = datetime.strptime(time_in_val, "%H:%M").time()
                         except ValueError:
                              pass # Keep default or None
            except Exception:
                pass # Keep default or None on error


            try:
                time_out_val = worksheet["D8"].value
                if isinstance(time_out_val, datetime.time):
                    data["time_out"] = time_out_val
                elif isinstance(time_out_val, datetime):
                    data["time_out"] = time_out_val.time()
                elif isinstance(time_out_val, str) and time_out_val:
                    try:
                        data["time_out"] = datetime.strptime(time_out_val, "%H:%M:%S").time()
                    except ValueError:
                         try:
                              data["time_out"] = datetime.strptime(time_out_val, "%H:%M").time()
                         except ValueError:
                              pass # Keep default or None
            except Exception:
                pass # Keep default or None on error

            try: data["period"] = worksheet["D4"].value
            except Exception: pass


            # Metadata from snippet 2 save locations
            try: data["observer_name"] = worksheet["AA1"].value
            except Exception: pass
            try: data["teacher_name"] = worksheet["AA2"].value
            except Exception: pass
            try: data["observation_type"] = worksheet["AA3"].value
            except Exception: pass
            # Timestamp AA4 - not loaded into input
            try: data["operator"] = worksheet["AA5"].value
            except Exception: pass
            try: data["school_name"] = worksheet["AA6"].value
            except Exception: pass
            try: data["overall_notes"] = worksheet["AA7"].value
            except Exception: pass
            try: data["teacher_email"] = worksheet["AA8"].value # Assuming AA8 for email
            except Exception: pass
            try: data["feedback_generated_timestamp"] = worksheet["AA9"].value # Assuming AA9 for feedback timestamp
            except Exception: pass


            # Rubric Scores and Notes - Read values saved in the sheet
            data["element_inputs"] = {} # Store inputs keyed by unique key like f"{domain}_{i}_rating/note"
            for idx, (domain, (start_cell, count)) in enumerate(rubric_structure.items()):
                col_rating = start_cell[0] # Column for rating (e.g., 'I')
                col_note = 'J' # Column for notes (based on snippet 2 save logic)
                try:
                    row_start = int(start_cell[1:])
                    for i in range(count):
                        row = row_start + i
                        rating_key = f"{domain}_{i}_rating"
                        note_key = f"{domain}_{i}_note"
                        # Read value from sheet, use try-except for individual cells
                        try:
                            rating_value_from_sheet = worksheet[f"{col_rating}{row}"].value
                            # Convert numeric ratings to int if they are floats (openpyxl might read ints as floats)
                            if isinstance(rating_value_from_sheet, float) and rating_value_from_sheet.is_integer():
                                rating_value_from_sheet = int(rating_value_from_sheet)
                             # Ensure "NA" is read correctly (handle case insensitive and potential None/empty)
                            elif isinstance(rating_value_from_sheet, str) and rating_value_from_sheet.strip().upper() == "NA":
                                 rating_value_from_sheet = "NA"
                             # Convert numbers read as text back to numbers if needed
                            elif isinstance(rating_value_from_sheet, str) and rating_value_from_sheet.strip().isdigit():
                                 rating_value_from_sheet = int(rating_value_from_sheet.strip())
                             # Handle empty cells read as None or empty strings, default to "NA"
                            elif rating_value_from_sheet is None or (isinstance(rating_value_from_sheet, str) and rating_value_from_sheet.strip() == ""):
                                 rating_value_from_sheet = "NA"


                            data["element_inputs"][rating_key] = rating_value_from_sheet
                        except Exception:
                            data["element_inputs"][rating_key] = "NA" # Default to NA on error

                        try:
                            note_value_from_sheet = worksheet[f"{col_note}{row}"].value
                            data["element_inputs"][note_key] = note_value_from_sheet if note_value_from_sheet is not None else ""
                        except Exception:
                            data["element_inputs"][note_key] = "" # Default to empty string on error

                except Exception as e:
                    # st.warning(f"Error loading rubric data for domain {domain}: {e}") # Avoid st.warning inside function
                    print(f"Error loading rubric data for domain {domain}: {e}")
                    # Continue to next domain even if one fails


            return data

        # --- Rubric Structure Definition (Moved here as it's needed for input fields) ---
        rubric_domains_structure = {
            "Domain 1": ("I11", 5), # Starting Cell (Rating Column), Number of Elements
            "Domain 2": ("I20", 3),
            "Domain 3": ("I27", 4),
            "Domain 4": ("I35", 3),
            "Domain 5": ("I42", 2),
            "Domain 6": ("I48", 2),
            "Domain 7": ("I54", 2),
            "Domain 8": ("I60", 3),
            "Domain 9": ("I67", 2)
        }

        # --- Rubric Descriptor Reading ---
        # Read descriptors from 'LO 1' template
        rubric_descriptors = {}
        try:
             template_ws = wb["LO 1"]
             # Assuming descriptor text for each element rating is in cells H11:H70 (approx)
             # Need to map element/rating to the descriptor text
             # Based on snippet 2, descriptors seem to be linked to specific ratings within the cells around the rating columns (H column)
             # Let's refine this based on a plausible structure:
             # Assuming cell H[row] contains descriptors for the element at row [row], possibly combined or for a default rating.
             # A better approach might be to map based on the template's layout:
             # Example: H11-H15 for Domain 1 elements 1-5, H11 might have descriptors for element 1.
             # To get specific descriptor per rating (1-6), the sheet likely has columns for each rating's descriptor.
             # Let's assume descriptor for rating R for element at row X is in cell Col(R+offset)X
             # This is complex to infer without the exact sheet structure.
             # As a simplified approach, let's try to read *all* text from the 'LO 1' sheet
             # around the rubric rating column (Column I in snippet 2) and try to associate it.
             # A more robust way requires knowing the *exact* cell layout of descriptors for each rating.

             # Let's try reading from the 'LO 1' sheet, focusing on columns near the ratings (H and potentially others)
             # Iterate through the rows covered by the rubric structure
             for domain, (start_cell, count) in rubric_domains_structure.items():
                  try:
                       row_start = int(start_cell[1:])
                       for i in range(count):
                           row = row_start + i
                           element_label = template_ws[f"B{row}"].value # Assuming Element Label is in Column B
                           if element_label:
                                element_key = f"{domain}_{i}"
                                # Attempt to read descriptor text from cells adjacent to the rating column (Column I)
                                # Let's assume Column H contains general descriptors or descriptors for rating 4/Good
                                # If descriptors for each rating are in columns E to J (ratings 1 to 6), we need that mapping.
                                # Based on template structure (implied by snippet 2), ratings are in I, notes in J.
                                # Let's assume descriptors for ratings 1-6 are in columns E through H of the rubric section.
                                # This is a guess and needs verification against the actual Excel file structure.
                                # Assuming descriptor for rating R for element at row X is in Column (E + R - 1) at row X
                                # Let's read descriptors for ratings 1-6 for each element row
                                rubric_descriptors[element_key] = {}
                                for rating_value in range(1, 7): # Ratings 1 through 6
                                     # Calculate column letter for rating R's descriptor (E=1, F=2, ..., J=6 if that's the pattern)
                                     # Assuming E is rating 1, F is 2, G is 3, H is 4, I is 5, J is 6
                                     # Looking at the template image snippet, the columns E to H seem to contain descriptor text.
                                     # Ratings dropdown implies 1-6. Let's assume E=1, F=2, G=3, H=4 and ratings 5 and 6 descriptors are also in H or another column.
                                     # A common pattern is E(1), F(2), G(3), H(4), I(5), J(6). However, I is used for ratings.
                                     # Let's assume descriptors for ratings 1-6 are in columns E, F, G, H, I, J or adjacent columns.
                                     # If ratings are I, notes J, maybe descriptors are E, F, G, H, K, L? This is too much guessing.
                                     # Let's revert to a simpler, less precise method if exact mapping isn't known:
                                     # Read ALL non-empty cells in columns E-H for each rubric row as potential descriptor text.
                                     # This won't give specific text per rating, but some context.
                                     descriptor_text_parts = []
                                     for col_idx in range(4, 8): # Columns E (index 4) to H (index 7)
                                          col_letter = get_column_letter(col_idx + 1)
                                          cell_value = template_ws[f"{col_letter}{row}"].value
                                          if cell_value is not None and isinstance(cell_value, str) and cell_value.strip():
                                               descriptor_text_parts.append(cell_value.strip())

                                     # Store joined text under a general key, not per rating, if exact mapping is complex
                                     rubric_descriptors[element_key]['general'] = " ".join(descriptor_text_parts) if descriptor_text_parts else strings["info_no_descriptors"]

                                     # A more accurate way, if we knew the mapping, would be:
                                     # col_desc_for_rating = get_column_letter(descriptor_base_col_index + rating_value - 1) # e.g., E=1, F=2...
                                     # descriptor_cell_value = template_ws[f"{col_desc_for_rating}{row}"].value
                                     # if descriptor_cell_value:
                                     #      rubric_descriptors[element_key][str(rating_value)] = str(descriptor_cell_value).strip()
                                     # else:
                                     #      rubric_descriptors[element_key][str(rating_value)] = strings["info_no_descriptors"]
                                     # This specific mapping is missing. Let's stick to reading columns E-H for now.

                  except Exception as e:
                       print(f"Error reading descriptors for domain {domain} element {i}: {e}")
                       # Add placeholder descriptor if reading fails
                       rubric_descriptors[f"{domain}_{i}"] = {'general': strings["info_no_descriptors"]}

        except KeyError:
            st.warning("Template sheet 'LO 1' not found. Cannot read rubric structure or descriptors.")
            # Set empty structure and descriptors if template is missing
            rubric_domains_structure = {}
            rubric_descriptors = {}
        except Exception as e:
            st.error(f"Error reading rubric structure/descriptors from template: {e}")
            # Set empty structure and descriptors on error
            rubric_domains_structure = {}
            rubric_descriptors = {}


        # --- Logic based on selected sheet/create new ---
        if selected_option == strings["option_create_new"]:
            # Determine the name for the new sheet
            next_index = 1
            existing_lo_numbers = []
            for sheet in wb.sheetnames:
                 if sheet.startswith("LO "):
                      suffix = sheet[3:].strip() # Get suffix and strip whitespace
                      if suffix.isdigit():
                           existing_lo_numbers.append(int(suffix))

            if existing_lo_numbers:
                 next_index = max(existing_lo_numbers) + 1

            sheet_name_to_process = f"LO {next_index}" # This is the *target* name for the new sheet
            is_new_sheet = True

            # Initialize session state for new sheet (except auth email which persists)
            st.session_state.update({
                 'target_sheet_name': sheet_name_to_process,
                 'observer_name': None,
                 'teacher_name': None,
                 'teacher_email': None, # Start fresh for teacher email on new sheet
                 'operator': None,
                 'school_name': None,
                 'grade': None,
                 'subject': None,
                 'gender': None,
                 'students': None,
                 'males': None,
                 'females': None,
                 'time_in': None,
                 'time_out': None,
                 'observation_date': datetime.now().date(), # Default to today's date
                 'period': None,
                 'observation_type': strings["option_individual"], # Default observation type
                 'overall_notes': None,
                 'checkbox_send_feedback': False,
                 'element_inputs': {} # Initialize empty dict for element scores/notes
            })

            # Initialize element inputs in session state with default "NA" rating and empty notes for the new sheet
            for idx, (domain, (start_cell, count)) in enumerate(rubric_domains_structure.items()):
                 for i in range(count):
                      rating_key = f"{domain}_{i}_rating"
                      note_key = f"{domain}_{i}_note"
                      st.session_state['element_inputs'][rating_key] = "NA" # Default rating
                      st.session_state['element_inputs'][note_key] = "" # Default note

            st.info(strings["subheader_filling_data"].format(sheet_name_to_process))
            ws_to_load_from = wb["LO 1"] # Rubric structure comes from the template

        else: # Selected an existing sheet
            sheet_name_to_process = selected_option
            is_new_sheet = False
            st.session_state['target_sheet_name'] = sheet_name_to_process # Store the selected sheet name

            try:
                ws_to_load_from = wb[sheet_name_to_process] # Get the selected sheet object
                st.subheader(strings["subheader_filling_data"].format(sheet_name_to_process))

                # Load existing data into session state from the selected sheet
                # Pass the rubric structure to load_existing_data to know where to look for element scores/notes
                existing_data = load_existing_data(ws_to_load_from, rubric_domains_structure)

                # Update session state with loaded data, preserving auth email
                for key, value in existing_data.items():
                     if key != 'auth_email_input': # Do not overwrite auth email from input box
                          st.session_state[key] = value

                 # Ensure element_inputs is initialized even if loading failed for it
                if 'element_inputs' not in st.session_state:
                     st.session_state['element_inputs'] = {}

            except KeyError:
                st.error(f"Error: Selected sheet '{sheet_name_to_process}' not found or could not be accessed.")
                # Reset sheet selector if sheet is missing
                st.session_state['sheet_selector'] = strings["option_create_new"] # Reset to 'Create new' by key
                st.rerun()
                st.stop()
            except Exception as e:
                st.error(f"Error loading data from sheet '{sheet_name_to_process}': {e}")
                # Reset sheet selector if loading fails
                st.session_state['sheet_selector'] = strings["option_create_new"] # Reset to 'Create new' by key
                st.rerun()
                st.stop()


        # --- Input Form (Re-enabled) ---
        # Ensure there's a valid sheet name to process before showing inputs
        if sheet_name_to_process:
            st.info(f"**Target Sheet:** `{sheet_name_to_process}`") # Indicate which sheet is being edited/created

            # Basic Information Inputs (using st.session_state for initial value and persistence)
            # Use default values from session state. If session state is None, widgets will use their own defaults.
            st.markdown("---")
            st.subheader("Basic Information")
            cols = st.columns(2)
            with cols[0]:
                st.session_state['observer_name'] = st.text_input(strings["label_observer_name"], value=st.session_state.get('observer_name', ''), key='observer_name_input')
                st.session_state['teacher_name'] = st.text_input(strings["label_teacher_name"], value=st.session_state.get('teacher_name', ''), key='teacher_name_input')
                st.session_state['teacher_email'] = st.text_input(strings["label_teacher_email"], value=st.session_state.get('teacher_email', ''), key='teacher_email_input')
                st.session_state['operator'] = st.text_input(strings["label_operator"], value=st.session_state.get('operator', ''), key='operator_input')
                st.session_state['school_name'] = st.text_input(strings["label_school_name"], value=st.session_state.get('school_name', ''), key='school_name_input')


            with cols[1]:
                st.session_state['grade'] = st.text_input(strings["label_grade"], value=st.session_state.get('grade', ''), key='grade_input')
                st.session_state['subject'] = st.text_input(strings["label_subject"], value=st.session_state.get('subject', ''), key='subject_input')
                st.session_state['gender'] = st.selectbox(strings["label_gender"], ["Male", "Female", "Mixed", ""], index=["Male", "Female", "Mixed", ""].index(st.session_state.get('gender', '') or ""), key='gender_input') # Handle None/empty string gracefully

                # Numeric inputs with validation hint
                st.session_state['students'] = st.number_input(strings["label_students"], min_value=0, value=st.session_state.get('students', 0) or 0, step=1, key='students_input', format="%d") # Use 0 as default if None
                st.session_state['males'] = st.number_input(strings["label_males"], min_value=0, value=st.session_state.get('males', 0) or 0, step=1, key='males_input', format="%d")
                st.session_state['females'] = st.number_input(strings["label_females"], min_value=0, value=st.session_state.get('females', 0) or 0, step=1, key='females_input', format="%d")

            cols_date_time = st.columns(3)
            with cols_date_time[0]:
                 # Ensure observation_date is a date object for the date_input widget
                 default_date = st.session_state.get('observation_date', datetime.now().date())
                 if isinstance(default_date, datetime):
                     default_date = default_date.date()
                 elif not isinstance(default_date, date):
                      default_date = datetime.now().date() # Fallback to today if neither
                 st.session_state['observation_date'] = st.date_input(strings["label_observation_date"], value=default_date, key='observation_date_input')

            with cols_date_time[1]:
                # Ensure time_in is a time object for the time_input widget
                default_time_in = st.session_state.get('time_in', None) # Keep None as default
                if isinstance(default_time_in, datetime): # Handle if loaded as datetime
                     default_time_in = default_time_in.time()
                st.session_state['time_in'] = st.time_input(strings["label_time_in"], value=default_time_in, key='time_in_input')

            with cols_date_time[2]:
                # Ensure time_out is a time object for the time_input widget
                default_time_out = st.session_state.get('time_out', None) # Keep None as default
                if isinstance(default_time_out, datetime): # Handle if loaded as datetime
                     default_time_out = default_time_out.time()
                st.session_state['time_out'] = st.time_input(strings["label_time_out"], value=default_time_out, key='time_out_input')

            # Calculate and display Lesson Duration
            lesson_duration_minutes = None
            duration_type = ""
            duration_display = strings["warning_calculate_duration"]

            if st.session_state.get('time_in') is not None and st.session_state.get('time_out') is not None:
                 # Convert time objects to datetime on an arbitrary date for comparison
                 try:
                      # Use today's date or observation date if available, otherwise arbitrary
                      arbitrary_date = st.session_state.get('observation_date', date.today())
                      if not isinstance(arbitrary_date, date): # Ensure it's a date object
                           arbitrary_date = date.today()

                      dt_in = datetime.combine(arbitrary_date, st.session_state['time_in'])
                      dt_out = datetime.combine(arbitrary_date, st.session_state['time_out'])

                      # Handle observations spanning midnight
                      if dt_out < dt_in:
                          dt_out += timedelta(days=1)

                      duration_seconds = (dt_out - dt_in).total_seconds()
                      lesson_duration_minutes = round(duration_seconds / 60) # Round to nearest minute

                      if lesson_duration_minutes >= 40: # Threshold for full lesson vs walkthrough
                          duration_type = strings["duration_full_lesson"]
                      else:
                          duration_type = strings["duration_walkthrough"]

                      duration_display = strings["label_lesson_duration"].format(lesson_duration_minutes, duration_type)

                 except Exception as e:
                      duration_display = strings["warning_could_not_calculate_duration"]
                      st.warning(f"Duration calculation error: {e}") # Show error for debugging


            st.info(duration_display) # Display duration info

            st.session_state['period'] = st.text_input(strings["label_period"], value=st.session_state.get('period', ''), key='period_input')
            st.session_state['observation_type'] = st.selectbox(strings["label_obs_type"], [strings["option_individual"], strings["option_joint"]], index=[strings["option_individual"], strings["option_joint"]].index(st.session_state.get('observation_type', strings["option_individual"]) or strings["option_individual"]), key='observation_type_input') # Handle None/empty string

            st.markdown("---")
            st.subheader(strings["subheader_rubric_scores"])

            # Rubric Inputs (using st.session_state for values)
            # Ratings options including "NA"
            rating_options = ["NA", 1, 2, 3, 4, 5, 6]

            # Load the template worksheet once to get element labels and potential descriptors
            template_ws = None
            try:
                template_ws = wb["LO 1"]
            except KeyError:
                st.error("Template sheet 'LO 1' not found. Cannot display rubric details.")
                template_ws = None # Ensure it's None if not found

            if template_ws and rubric_domains_structure:
                for idx, (domain, (start_cell, count)) in enumerate(rubric_domains_structure.items()):
                     st.markdown(f"#### {domain}: {template_ws[f'B{int(start_cell[1:]) - 1}'].value if template_ws and f'B{int(start_cell[1:]) - 1}' in template_ws else domain}") # Display Domain Title from template (adjust row if needed)
                     # Optional: Add expander for domain-level guidance if available in the template
                     # st.expander(f"Guidance for {domain}").markdown("Guidance text goes here...")

                     for i in range(count):
                          row = int(start_cell[1:]) + i
                          element_label = template_ws[f"B{row}"].value if template_ws and f"B{row}" in template_ws else f"Element {i+1}"
                          element_key = f"{domain}_{i}"

                          st.markdown(f"##### {element_label}")

                          # Get descriptor text for this element (using the simplified 'general' lookup)
                          descriptor_text = rubric_descriptors.get(element_key, {}).get('general', strings["info_no_descriptors"])
                          st.expander(strings["expander_rubric_descriptors"]).markdown(descriptor_text) # Display descriptor in expander

                          cols_rating_note = st.columns(2)
                          with cols_rating_note[0]:
                             rating_key = f"{domain}_{i}_rating"
                             # Ensure the default value is one of the options, fallback to "NA"
                             current_rating = st.session_state['element_inputs'].get(rating_key, "NA")
                             if current_rating not in rating_options:
                                 current_rating = "NA"
                             st.session_state['element_inputs'][rating_key] = st.selectbox(
                                 strings["label_rating_for"].format(element_label),
                                 rating_options,
                                 index=rating_options.index(current_rating), # Set index based on current value
                                 key=rating_key # Unique key for session state
                             )
                          with cols_rating_note[1]:
                             note_key = f"{domain}_{i}_note"
                             st.session_state['element_inputs'][note_key] = st.text_area(
                                 strings["label_write_notes"].format(element_label),
                                 value=st.session_state['element_inputs'].get(note_key, ""),
                                 key=note_key # Unique key for session state
                             )

            else:
                 st.info("Rubric structure could not be loaded from the template sheet.")


            st.markdown("---")
            st.session_state['overall_notes'] = st.text_area(strings["label_overall_notes"], value=st.session_state.get('overall_notes', ''), key='overall_notes_input')

            st.markdown("---")
            # Checkbox to generate feedback report/PDF
            st.session_state['checkbox_send_feedback'] = st.checkbox(strings["checkbox_send_feedback"], value=st.session_state.get('checkbox_send_feedback', False), key='send_feedback_checkbox')


            # --- Save Button ---
            if st.button(strings["button_save_observation"], key='save_observation_button'):
                 # --- Validation ---
                 # Check essential basic info fields are filled
                 essential_fields = {
                     strings["label_observer_name"]: st.session_state.get('observer_name'),
                     strings["label_teacher_name"]: st.session_state.get('teacher_name'),
                     strings["label_school_name"]: st.session_state.get('school_name'),
                     strings["label_grade"]: st.session_state.get('grade'),
                     strings["label_subject"]: st.session_state.get('subject'),
                     strings["label_gender"]: st.session_state.get('gender'),
                     strings["label_observation_date"]: st.session_state.get('observation_date'),
                 }
                 missing_essential = [label for label, value in essential_fields.items() if value is None or (isinstance(value, str) and value.strip() == "")]

                 if missing_essential:
                     st.warning(strings["warning_fill_essential"])
                     st.stop() # Stop execution if essential fields are missing

                 # Check numeric fields are valid numbers if filled
                 numeric_fields = {
                      strings["label_students"]: st.session_state.get('students'),
                      strings["label_males"]: st.session_state.get('males'),
                      strings["label_females"]: st.session_state.get('females'),
                 }
                 invalid_numeric = [label for label, value in numeric_fields.items() if value is not None and not isinstance(value, (int, float))]

                 if invalid_numeric:
                      st.warning(strings["warning_numeric_fields"])
                      st.stop() # Stop execution if numeric fields are invalid

                 # Check at least one rubric score is entered (not all "NA")
                 all_ratings = [st.session_state['element_inputs'].get(f"{domain}_{i}_rating") for domain, (start_cell, count) in rubric_domains_structure.items() for i in range(count)]
                 if all(r is None or (isinstance(r, str) and r.upper() == "NA") for r in all_ratings):
                     st.warning("Please enter at least one rubric rating.")
                     st.stop() # Stop execution

                 # --- Get/Create Target Sheet ---
                 target_sheet_name = st.session_state['target_sheet_name']
                 ws = None

                 if is_new_sheet:
                     try:
                         # Copy the template sheet ("LO 1") to the new sheet name
                         template_ws = wb["LO 1"]
                         ws = wb.copy_worksheet(template_ws)
                         ws.title = target_sheet_name # Rename the copied sheet
                         st.success(strings["success_sheet_created"].format(target_sheet_name))
                     except KeyError:
                         st.error(strings["error_template_not_found"])
                         st.stop() # Cannot create new sheet without template
                     except Exception as e:
                         st.error(f"Error creating new sheet '{target_sheet_name}': {e}")
                         st.stop()
                 else:
                     try:
                         ws = wb[target_sheet_name]
                     except KeyError:
                          st.error(f"Error: Target sheet '{target_sheet_name}' not found during save.")
                          st.stop() # Should not happen if sheet selector logic is correct, but safety check


                 # --- Write Data to Worksheet ---
                 # Basic Info - Referencing cell locations from snippet 2 save logic
                 try:
                     ws["AA1"].value = st.session_state.get('observer_name')
                     ws["AA2"].value = st.session_state.get('teacher_name')
                     ws["AA3"].value = st.session_state.get('observation_type')
                     ws["AA4"].value = datetime.now() # Timestamp of save
                     ws["AA5"].value = st.session_state.get('operator')
                     ws["AA6"].value = st.session_state.get('school_name')
                     ws["AA7"].value = st.session_state.get('overall_notes') # Overall notes
                     ws["AA8"].value = st.session_state.get('teacher_email') # Teacher email

                     # Basic Lesson Details
                     ws["B1"].value = st.session_state.get('grade') # Assuming B1 for Grade
                     ws["B5"].value = st.session_state.get('gender')
                     ws["B6"].value = st.session_state.get('students')
                     ws["B7"].value = st.session_state.get('males')
                     ws["B8"].value = st.session_state.get('females')
                     ws["D2"].value = st.session_state.get('subject')
                     ws["D4"].value = st.session_state.get('period')
                     ws["D7"].value = st.session_state.get('time_in')
                     ws["D8"].value = st.session_state.get('time_out')
                     # D10 for Observation Date - Ensure it's written as a date object
                     obs_date = st.session_state.get('observation_date')
                     if obs_date and isinstance(obs_date, date):
                          ws["D10"].value = obs_date
                     else:
                          ws["D10"].value = None # Clear if invalid


                 except Exception as e:
                      st.error(f"Error writing basic information to sheet: {e}")
                      # Continue trying to save other data, but alert the user


                 # Rubric Scores and Notes - Write values from session state
                 try:
                     for domain, (start_cell, count) in rubric_domains_structure.items():
                         col_rating = start_cell[0]
                         col_note = 'J' # Notes column
                         row_start = int(start_cell[1:])
                         for i in range(count):
                             row = row_start + i
                             rating_key = f"{domain}_{i}_rating"
                             note_key = f"{domain}_{i}_note"

                             # Write rating, handling "NA" vs numbers
                             rating_value = st.session_state['element_inputs'].get(rating_key, "NA")
                             ws[f"{col_rating}{row}"].value = rating_value if rating_value != "NA" else "NA" # Save "NA" as string

                             # Write note
                             note_value = st.session_state['element_inputs'].get(note_key, "")
                             ws[f"{col_note}{row}"].value = note_value if note_value is not None else ""


                 except Exception as e:
                     st.error(f"Error writing rubric data to sheet: {e}")
                     # Continue, but alert the user

                 # --- Calculate Overall Score and Domain Averages ---
                 # Calculate from the saved data on the worksheet or from session state?
                 # Calculating from session state is faster as data is already in memory.
                 # Need the domain structure mapping again.
                 domain_data_for_feedback = {}
                 overall_scores_list = []

                 for domain_name, (start_cell, count) in rubric_domains_structure.items():
                      domain_ratings = []
                      elements_data = []
                      for i in range(count):
                          rating_key = f"{domain_name}_{i}_rating"
                          note_key = f"{domain_name}_{i}_note"
                          element_label = f"Element {i+1}" # Default if not found in template_ws
                          if template_ws:
                              try:
                                  label_row = int(start_cell[1:]) + i
                                  label_from_sheet = template_ws[f"B{label_row}"].value
                                  if label_from_sheet:
                                       element_label = label_from_sheet
                              except Exception:
                                   pass # Use default label if reading from template fails


                          rating = st.session_state['element_inputs'].get(rating_key)
                          note = st.session_state['element_inputs'].get(note_key, "")

                          # Add numeric ratings to list for averaging, ignore "NA"
                          if isinstance(rating, (int, float)) and not math.isnan(rating):
                               domain_ratings.append(float(rating))
                               overall_scores_list.append(float(rating)) # Add to overall list

                          # Get descriptor for this element and rating (using simplified general descriptor)
                          element_key_for_descriptor = f"{domain_name}_{i}"
                          descriptor_text = rubric_descriptors.get(element_key_for_descriptor, {}).get('general', strings["info_no_descriptors"])

                          elements_data.append({
                             "label": element_label,
                             "rating": rating if rating is not None else "NA",
                             "note": note,
                             "descriptor": descriptor_text # Pass the descriptor text
                          })


                      domain_average = statistics.mean(domain_ratings) if domain_ratings else np.nan # Use np.nan for no scores
                      domain_judgment = get_performance_level(domain_average, strings)

                      domain_title = domain_name # Default domain title
                      if template_ws:
                           try:
                                # Assuming domain title is above the first element, adjust row accordingly
                                title_row = int(start_cell[1:]) - 1
                                title_from_sheet = template_ws[f"B{title_row}"].value
                                if title_from_sheet:
                                     domain_title = title_from_sheet
                           except Exception:
                                pass # Use default domain name

                      domain_data_for_feedback[domain_name] = {
                          "title": domain_title,
                          "average": domain_average,
                          "judgment": domain_judgment,
                          "elements": elements_data
                      }

                 overall_average_score = statistics.mean(overall_scores_list) if overall_scores_list else np.nan # Use np.nan for no scores
                 overall_judgment = get_performance_level(overall_average_score, strings)

                 # --- Update Feedback Log Sheet ---
                 feedback_log_sheet_name = strings["feedback_log_sheet_name"]
                 try:
                     if feedback_log_sheet_name not in wb.sheetnames:
                         # Create log sheet if it doesn't exist and add headers
                         log_ws = wb.create_sheet(feedback_log_sheet_name)
                         log_ws.append(strings["feedback_log_header"]) # Add header row
                     else:
                         log_ws = wb[feedback_log_sheet_name]
                         # Find the next empty row
                         next_row = log_ws.max_row + 1

                     # Append a new row with observation summary data
                     log_data_row = [
                         target_sheet_name,
                         st.session_state.get('observer_name'),
                         st.session_state.get('teacher_name'),
                         st.session_state.get('teacher_email'),
                         st.session_state.get('school_name'),
                         st.session_state.get('subject'),
                         st.session_state.get('observation_date'), # Date object should be okay
                         overall_judgment, # Overall judgment string
                         overall_average_score if not np.isnan(overall_average_score) else None, # Save numeric score or None
                         st.session_state.get('overall_notes') # Include overall notes in the log
                     ]
                     log_ws.append(log_data_row)
                     st.session_state['feedback_generated_timestamp'] = datetime.now() # Record timestamp of feedback generation
                     st.success(strings["success_feedback_log_updated"])

                 except Exception as e:
                     st.error(strings["error_updating_log"].format(e))
                     # Continue trying to save the workbook even if log update failed


                 # --- Generate Feedback Report Text ---
                 feedback_content_text = ""
                 if st.session_state.get('checkbox_send_feedback', False):
                     try:
                         teacher_name = st.session_state.get('teacher_name', 'Teacher')
                         obs_date_display = st.session_state.get('observation_date')
                         if isinstance(obs_date_display, (date, datetime)):
                              obs_date_display = obs_date_display.strftime("%Y-%m-%d")
                         else:
                              obs_date_display = "N/A"

                         feedback_content_text += strings["feedback_greeting"].format(teacher_name, obs_date_display)
                         feedback_content_text += strings["feedback_observer"].format(st.session_state.get('observer_name', 'N/A'))
                         feedback_content_text += strings["feedback_duration"].format(f"{lesson_duration_minutes} minutes ({duration_type})" if lesson_duration_minutes is not None else 'N/A')
                         feedback_content_text += strings["feedback_subject_fb"].format(st.session_state.get('subject', 'N/A'))
                         feedback_content_text += strings["feedback_school"].format(st.session_state.get('school_name', 'N/A'))

                         feedback_content_text += strings["feedback_summary_header"]

                         # Summarize scores per domain
                         if domain_data_for_feedback:
                              for domain_name, domain_info in domain_data_for_feedback.items():
                                   feedback_content_text += strings["feedback_domain_header"].format(domain_name, domain_info.get("title", domain_name))
                                   for element in domain_info.get("elements", []):
                                         feedback_content_text += strings["feedback_element_rating"].format(element["label"], element["rating"])
                                         # Add descriptor text if available
                                         if element["descriptor"] and element["descriptor"] != strings["info_no_descriptors"]:
                                               feedback_content_text += strings["feedback_descriptor_for_rating"].format(element["rating"], element["descriptor"])
                                         # Add note if available
                                         if element["note"] and element["note"].strip():
                                              feedback_content_text += f"  *Notes:* {element['note'].strip()}\n"

                                   if domain_info.get("average") is not None and not np.isnan(domain_info["average"]):
                                        feedback_content_text += strings["feedback_domain_average"].format(domain_info["average"])
                                   feedback_content_text += "\n" # Add a blank line between domains

                         # Add overall score
                         if overall_average_score is not None and not np.isnan(overall_average_score):
                             feedback_content_text += strings["feedback_overall_score"].format(overall_average_score)

                         # Add Performance Summary and Plan/Next Steps based on overall judgment
                         feedback_content_text += strings["feedback_performance_summary"]
                         feedback_content_text += strings["overall_performance_level_text"].format(overall_judgment) + "\n\n"

                         # Add general plan/steps based on overall judgment
                         if overall_judgment == strings["perf_level_very_weak"]:
                              feedback_content_text += strings["feedback_support_plan_intro"] + strings["plan_very_weak_overall"] + "\n\n"
                         elif overall_judgment == strings["perf_level_weak"]:
                              feedback_content_text += strings["feedback_support_plan_intro"] + strings["plan_weak_overall"] + "\n\n"
                              # Could add domain-specific plans here if domain judgment is weak (example logic)
                              for domain_name, domain_info in domain_data_for_feedback.items():
                                   if domain_info.get("judgment") == strings["perf_level_weak"]:
                                        # Need domain-specific guidance/actions here, currently placeholders in strings
                                        feedback_content_text += strings["plan_weak_domain"].format(domain_info.get("title", domain_name), "...") + "\n" # Add details based on domain
                                        feedback_content_text += "[Provide specific actionable steps here]\n\n" # Add more specific steps

                         elif overall_judgment == strings["perf_level_acceptable"]:
                              feedback_content_text += strings["feedback_next_steps_intro"] + strings["steps_acceptable_overall"] + "\n\n"
                         elif overall_judgment == strings["perf_level_good"]:
                             feedback_content_text += strings["feedback_next_steps_intro"] + strings["steps_good_overall"] + "\n\n"
                             # Could add domain-specific steps here if domain judgment is good/excellent
                             for domain_name, domain_info in domain_data_for_feedback.items():
                                   if domain_info.get("judgment") == strings["perf_level_good"]:
                                        # Need domain-specific guidance/actions here, currently placeholders in strings
                                        feedback_content_text += strings["steps_good_domain"].format(domain_info.get("title", domain_name), "...") + "\n" # Add details based on domain
                                        feedback_content_text += "[Provide specific growth opportunities here]\n\n" # Add more specific steps

                         elif overall_judgment == strings["perf_level_excellent"]:
                              feedback_content_text += strings["feedback_next_steps_intro"] + strings["steps_excellent_overall"] + "\n\n"
                              for domain_name, domain_info in domain_data_for_feedback.items():
                                   if domain_info.get("judgment") == strings["perf_level_excellent"]:
                                        feedback_content_text += strings["steps_excellent_domain"].format(domain_info.get("title", domain_name), "...") + "\n" # Add details based on domain
                                        feedback_content_text += "[Provide opportunities for leadership/sharing here]\n\n" # Add more specific opportunities
                         else: # Handles N/A case or others
                             feedback_content_text += strings["no_specific_plan_needed"] + "\n\n"


                         feedback_content_text += strings["feedback_closing"]
                         feedback_content_text += strings["feedback_regards"]

                         # Store feedback text in session state (optional, useful for display/debugging)
                         st.session_state['generated_feedback_text'] = feedback_content_text
                         st.success(strings["success_feedback_generated"])

                     except Exception as e:
                          st.error(f"Error generating feedback text: {e}")
                          st.session_state['generated_feedback_text'] = "Error generating feedback text."
                          feedback_content_text = "" # Clear text if generation failed

                 # --- Generate PDF (if checkbox is checked) ---
                 pdf_buffer = None
                 if st.session_state.get('checkbox_send_feedback', False) and feedback_content_text:
                      st.info("Generating PDF feedback report...")
                      # Pass all necessary data to the PDF function
                      pdf_data = {
                          "observer_name": st.session_state.get('observer_name'),
                          "teacher_name": st.session_state.get('teacher_name'),
                          "teacher_email": st.session_state.get('teacher_email'),
                          "operator": st.session_state.get('operator'),
                          "school_name": st.session_state.get('school_name'),
                          "grade": st.session_state.get('grade'),
                          "subject": st.session_state.get('subject'),
                          "gender": st.session_state.get('gender'),
                          "students": st.session_state.get('students'),
                          "males": st.session_state.get('males'),
                          "females": st.session_state.get('females'),
                          "observation_date": st.session_state.get('observation_date'),
                          "time_in": st.session_state.get('time_in'),
                          "time_out": st.session_state.get('time_out'),
                          "duration_display": f"{lesson_duration_minutes} minutes ({duration_type})" if lesson_duration_minutes is not None else strings["warning_could_not_calculate_duration"],
                          "period": st.session_state.get('period'),
                          "observation_type": st.session_state.get('observation_type'),
                          "overall_notes": st.session_state.get('overall_notes'),
                          "overall_score_display": strings["overall_score_value"].format(overall_average_score) if not np.isnan(overall_average_score) else strings["overall_score_na"],
                          "domain_data": domain_data_for_feedback, # Pass the structured domain data
                      }
                      pdf_buffer = generate_observation_pdf(pdf_data, feedback_content_text, strings)

                      if pdf_buffer:
                           st.success(strings["success_pdf_generated"])
                      else:
                           st.error("Failed to generate PDF feedback report.")


                 # --- Save Workbook to File ---
                 try:
                      wb.save(DEFAULT_FILE)
                      st.success(strings["success_data_saved"])
                      # Reload workbook into session state after saving
                      st.session_state.workbook = load_workbook(DEFAULT_FILE)
                      wb = st.session_state.workbook # Update the local wb variable
                 except Exception as e:
                      st.error(strings["error_saving_workbook"].format(e))


                 # --- Provide Download Buttons ---
                 st.markdown("---")
                 st.markdown("##### Download Options")
                 col1, col2, col3 = st.columns(3)

                 with col1:
                     # Download updated workbook
                     workbook_buffer = io.BytesIO()
                     wb.save(workbook_buffer)
                     workbook_buffer.seek(0)
                     st.download_button(
                         label=strings["download_workbook"],
                         data=workbook_buffer,
                         file_name=DEFAULT_FILE,
                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                     )

                 with col2:
                     # Download Feedback PDF (if generated)
                     if pdf_buffer:
                         teacher_name_for_pdf = st.session_state.get('teacher_name', 'feedback').replace(' ', '_').lower()
                         obs_date_for_pdf = st.session_state.get('observation_date')
                         date_str_for_pdf = obs_date_for_pdf.strftime("%Y%m%d") if isinstance(obs_date_for_pdf, (date, datetime)) else "undated"
                         pdf_file_name = f"observation_feedback_{teacher_name_for_pdf}_{date_str_for_pdf}.pdf"

                         st.download_button(
                             label=strings["download_feedback_pdf"],
                             data=pdf_buffer,
                             file_name=pdf_file_name,
                             mime="application/pdf"
                         )
                     else:
                          st.button(strings["download_feedback_pdf"], disabled=True, help="Generate feedback first") # Disabled button if no PDF

                 with col3:
                     # Download Feedback Log CSV (if log sheet exists)
                     feedback_log_sheet_name = strings["feedback_log_sheet_name"]
                     if feedback_log_sheet_name in wb.sheetnames:
                         try:
                             log_ws = wb[feedback_log_sheet_name]
                             # Read all data including headers
                             log_data = []
                             for row in log_ws.iter_rows(values_only=True):
                                  log_data.append(row)

                             if log_data:
                                 csv_buffer_log = io.StringIO()
                                 writer = csv.writer(csv_buffer_log)
                                 writer.writerows(log_data)
                                 csv_buffer_log.seek(0)

                                 st.download_button(
                                     label=strings["download_feedback_log_csv"],
                                     data=csv_buffer_log.getvalue(),
                                     file_name="feedback_log.csv",
                                     mime="text/csv"
                                 )
                             else:
                                  st.button(strings["download_feedback_log_csv"], disabled=True, help="Feedback log is empty.")
                         except Exception as e:
                             st.error(strings["error_generating_log_csv"].format(e))
                             st.button(strings["download_feedback_log_csv"], disabled=True, help="Error generating log CSV.")
                     else:
                         st.button(strings["download_feedback_log_csv"], disabled=True, help="Feedback Log sheet not found.")


                 # Optional: Display generated feedback text
                 if st.session_state.get('checkbox_send_feedback', False) and st.session_state.get('generated_feedback_text'):
                      st.markdown("---")
                      st.subheader("Generated Feedback Text")
                      st.text(st.session_state['generated_feedback_text']) # Use st.text for preformatted text


        else: # Fallback if sheet_name_to_process is None (e.g., template missing)
             st.warning(strings["warning_select_create_sheet"])


    # <--- End of Lesson Observation Input Page Logic --->


    elif page == strings["page_analytics"]:
        st.title(strings["title_analytics"])

        # Check if workbook is loaded before proceeding with analytics
        if not wb:
            st.warning(strings["warning_no_lo_sheets_analytics"]) # Reusing this string as appropriate
            st.stop()

        # --- Analytics Logic ---
        # Find all LO sheets (excluding the template) and the Feedback Log
        lo_sheets_data_list = [] # Use a list to build data before creating DataFrame
        feedback_log_data = pd.DataFrame()
        feedback_log_sheet_name = strings["feedback_log_sheet_name"]

        # Structure defining domains and their average cells in the LO sheets
        rubric_domains_avg_cells = {
            "Avg Domain 1": "I16", "Avg Domain 2": "I23", "Avg Domain 3": "I31",
            "Avg Domain 4": "I38", "Avg Domain 5": "I44", "Avg Domain 6": "I50",
            "Avg Domain 7": "I56", "Avg Domain 8": "I63", "Avg Domain 9": "I69"
        }

        try:
            # Load data from LO sheets
            lo_sheets_to_process = [sheet for sheet in wb.sheetnames if sheet.startswith("LO ") and sheet != "LO 1"]

            if not lo_sheets_to_process:
                 st.info(strings["warning_no_lo_sheets_analytics"])
            else:
                 for sheet_name in lo_sheets_to_process:
                     try:
                         ws = wb[sheet_name]
                         # Attempt to extract data points based on known cell locations
                         data = {
                             "Sheet": sheet_name,
                             # Use .value safely with checks
                             "Observer": ws["AA1"].value if "AA1" in ws and ws["AA1"].value is not None else None,
                             "Teacher": ws["AA2"].value if "AA2" in ws and ws["AA2"].value is not None else None,
                             "Operator": ws["AA5"].value if "AA5" in ws and ws["AA5"].value is not None else None,
                             "School": ws["AA6"].value if "AA6" in ws and ws["AA6"].value is not None else None,
                             "Grade": ws["B1"].value if "B1" in ws and ws["B1"].value is not None else None, # Assuming Grade is in B1 based on template layout
                             "Subject": ws["D2"].value if "D2" in ws and ws["D2"].value is not None else None,
                             "Gender": ws["B5"].value if "B5" in ws and ws["B5"].value is not None else None,
                             # Convert numbers to numeric directly, coerce errors
                             "Students": pd.to_numeric(ws["B6"].value if "B6" in ws else None, errors='coerce'),
                             "Males": pd.to_numeric(ws["B7"].value if "B7" in ws else None, errors='coerce'),
                             "Females": pd.to_numeric(ws["B8"].value if "B8" in ws else None, errors='coerce'),
                             "Observation Date": ws["D10"].value if "D10" in ws else None, # Assuming date is D10
                             "Observation Type": ws["AA3"].value if "AA3" in ws and ws["AA3"].value is not None else None,
                             "Overall Score": None, # Placeholder, calculated next
                             "Overall Judgment": None, # Placeholder, calculated next
                         }

                         # Extract Domain Averages from LO sheet cells
                         for domain_key, cell_ref in rubric_domains_avg_cells.items():
                             try:
                                 # Check if cell exists before accessing value
                                 avg_value = ws[cell_ref].value if cell_ref in ws else None
                                 # Convert to numeric, errors='coerce' turns non-numbers into NaN
                                 data[domain_key] = pd.to_numeric(avg_value, errors='coerce')
                             except Exception:
                                 data[domain_key] = pd.NA # Store pandas NA on error


                         # Calculate Overall Score and Judgment from Excel formulas if possible
                         # Assuming Overall Score is calculated somewhere, e.g., AM1
                         try:
                             # Check if cell exists before accessing value
                             overall_score_excel = ws["AM1"].value if "AM1" in ws else None
                             if isinstance(overall_score_excel, (int, float)) and not np.isnan(overall_score_excel):
                                 data["Overall Score"] = float(overall_score_excel)
                                 # Recalculate judgment based on the numeric score using the function
                                 data["Overall Judgment"] = get_performance_level(data["Overall Score"], strings)
                             else:
                                 data["Overall Score"] = None # Use None or pd.NA for missing/invalid
                                 data["Overall Judgment"] = strings["overall_score_na"]
                         except Exception:
                             data["Overall Score"] = None # Use None or pd.NA on error
                             data["Overall Judgment"] = strings["overall_score_na"]


                         # Append the data dictionary to the list
                         lo_sheets_data_list.append(data)

                     except Exception as e:
                         st.warning(f"Could not load data from sheet '{sheet_name}' for analytics: {e}")

            # Load data from Feedback Log sheet if it exists
            if feedback_log_sheet_name in wb.sheetnames:
                try:
                     log_ws = wb[feedback_log_sheet_name]
                     # Read data into a pandas DataFrame
                     # Assuming the log sheet has headers in the first row
                     data_rows = list(log_ws.iter_rows(min_row=2, values_only=True))
                     headers = [cell.value for cell in log_ws[1]] # Get headers from the first row

                     if headers and data_rows:
                          # Filter out empty rows if any
                          cleaned_data_rows = [row for row in data_rows if any(cell is not None and str(cell).strip() != "" for cell in row)]
                          if cleaned_data_rows:
                              feedback_log_data = pd.DataFrame(cleaned_data_rows, columns=headers)
                              # Attempt to convert 'Overall Score' to numeric, coercing errors
                              if 'Overall Score' in feedback_log_data.columns:
                                   feedback_log_data['Overall Score'] = pd.to_numeric(feedback_log_data['Overall Score'], errors='coerce')
                              # Convert 'Date' column to datetime if it exists and is not already
                              if 'Date' in feedback_log_data.columns:
                                   # Convert excel dates (numbers) or strings to datetime
                                   feedback_log_data['Date'] = pd.to_datetime(feedback_log_data['Date'], errors='coerce')
                    # No else needed here, feedback_log_data remains empty if no valid data
                except Exception as e:
                    st.error(f"Error loading {feedback_log_sheet_name}: {e}")

            # Convert extracted LO sheets data list to DataFrame
            all_obs_data = pd.DataFrame(lo_sheets_data_list)

        except Exception as e:
            st.error(strings["error_loading_analytics"].format(e))
            all_obs_data = pd.DataFrame() # Ensure DataFrame is empty on error


        if not all_obs_data.empty:
             # Convert 'Observation Date' column to datetime objects for robust comparison
             if 'Observation Date' in all_obs_data.columns:
                 all_obs_data['Observation Date'] = pd.to_datetime(all_obs_data['Observation Date'], errors='coerce')

             # Ensure numeric columns are numeric for calculations and filtering
             numeric_cols = ['Students', 'Males', 'Females'] + list(rubric_domains_avg_cells.keys()) + ['Overall Score']
             for col in numeric_cols:
                  if col in all_obs_data.columns:
                       all_obs_data[col] = pd.to_numeric(all_obs_data[col], errors='coerce')


             st.subheader(strings["subheader_data_summary"])
             st.write(f"Total Observations: {len(all_obs_data)}")
             if 'Overall Score' in all_obs_data.columns and not all_obs_data['Overall Score'].isna().all():
                 avg_overall_score = all_obs_data['Overall Score'].mean()
                 st.write(f"Overall Average Score: {avg_overall_score:.2f}")
             else:
                 st.write("Overall Average Score: N/A (No valid scores found)")


             # --- Filtering Options ---
             st.markdown("---")
             st.subheader(strings["subheader_filter_analyze"])

             # Get unique values for filters from ALL loaded LO sheets, drop NaNs, convert to list
             # Handle cases where a column might not exist in all_obs_data
             all_operators = sorted(all_obs_data['Operator'].dropna().unique().tolist()) if 'Operator' in all_obs_data.columns else []
             all_schools = sorted(all_obs_data['School'].dropna().unique().tolist()) if 'School' in all_obs_data.columns else []
             all_grades = sorted(all_obs_data['Grade'].dropna().unique().tolist()) if 'Grade' in all_obs_data.columns else []
             all_subjects = sorted(all_obs_data['Subject'].dropna().unique().tolist()) if 'Subject' in all_obs_data.columns else []
             all_teachers = sorted(all_obs_data['Teacher'].dropna().unique().tolist()) if 'Teacher' in all_obs_data.columns else []
             all_observers = sorted(all_obs_data['Observer'].dropna().unique().tolist()) if 'Observer' in all_obs_data.columns else []


             # Add filters
             filter_operator = st.selectbox(strings["filter_by_operator"], [strings["option_all"]] + all_operators)
             filter_school = st.selectbox(strings["filter_by_school"], [strings["option_all"]] + all_schools)
             filter_grade = st.selectbox(strings["filter_by_grade"], [strings["option_all"]] + all_grades)
             filter_subject = st.selectbox(strings["filter_by_subject"], [strings["option_all"]] + all_subjects)
             filter_teacher = st.selectbox(strings["filter_teacher"], [strings["option_all"]] + all_teachers)
             filter_observer = st.selectbox(strings["filter_by_observer_an"], [strings["option_all"]] + all_observers)


             # Date filtering
             st.markdown("##### Filter by Date")
             today = datetime.now().date()
             # Determine min/max dates from the loaded data, handling NaT values
             valid_dates = all_obs_data['Observation Date'].dropna() if 'Observation Date' in all_obs_data.columns else pd.Series(dtype='datetime64[ns]')

             # Safely get min/max date, fall back to today +/- 1 year if no valid dates
             min_date_data = valid_dates.min().date() if not valid_dates.empty else today - timedelta(days=365)
             max_date_data = valid_dates.max().date() if not valid_dates.empty else today + timedelta(days=7)

             # Ensure min/max dates are valid date objects for the widget
             min_date_input = min_date_data if isinstance(min_date_data, date) else today - timedelta(days=365)
             max_date_input = max_date_data if isinstance(max_date_data, date) else today + timedelta(days=7)

             # Adjust default start date to be no earlier than the min data date, and no later than today
             default_start_date = max(min_date_input, today - timedelta(days=365))
             default_start_date = min(default_start_date, today) # Cannot start filter in the future default

             # Adjust default end date to be no earlier than the max data date, and no earlier than today
             default_end_date = max(max_date_input, today)
             default_end_date = max(default_end_date, default_start_date) # End date cannot be before start date default


             try:
                  start_date = st.date_input(strings["filter_start_date"], value=default_start_date, min_value=min_date_input, max_value=max_date_input)
             except Exception:
                  start_date = st.date_input(strings["filter_start_date"], value=today - timedelta(days=365)) # Fallback default

             try:
                  end_date = st.date_input(strings["filter_end_date"], value=default_end_date, min_value=min_date_input, max_value=max_date_input)
             except Exception:
                  end_date = st.date_input(strings["filter_end_date"], value=today + timedelta(days=7)) # Fallback default


             # Apply Filters
             # Start with the full dataset
             filtered_data = all_obs_data.copy()

             # Apply categorical filters first, handling potential missing columns
             if filter_operator != strings["option_all"] and 'Operator' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['Operator'] == filter_operator].copy()

             if filter_school != strings["option_all"] and 'School' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['School'] == filter_school].copy()

             if filter_grade != strings["option_all"] and 'Grade' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['Grade'] == filter_grade].copy()

             if filter_subject != strings["option_all"] and 'Subject' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['Subject'] == filter_subject].copy()

             if filter_teacher != strings["option_all"] and 'Teacher' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['Teacher'] == filter_teacher].copy()

             if filter_observer != strings["option_all"] and 'Observer' in filtered_data.columns:
                  filtered_data = filtered_data[filtered_data['Observer'] == filter_observer].copy()

             # Apply date filter last, ensuring the column exists and has valid datetimes
             if 'Observation Date' in filtered_data.columns and not filtered_data['Observation Date'].isna().all():
                  # Filter out NaT values before comparison and convert to date for comparison with date pickers
                  filtered_data = filtered_data.dropna(subset=['Observation Date']).copy()
                  # Convert date picker results to pandas Timestamps for direct comparison with datetime64[ns]
                  start_timestamp = pd.Timestamp(start_date)
                  end_timestamp = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1) # Include the whole end day

                  filtered_data = filtered_data[(filtered_data['Observation Date'] >= start_timestamp) & (filtered_data['Observation Date'] <= end_timestamp)].copy()


             st.markdown("---")
             st.subheader(strings["subheader_avg_score_filtered"])

             if not filtered_data.empty:
                 # Calculate average overall score for filtered data, ignoring NaNs
                 if 'Overall Score' in filtered_data.columns and not filtered_data['Overall Score'].isna().all():
                      avg_filtered_score = filtered_data['Overall Score'].mean()
                      st.write(f"Average Overall Score for Filtered Data: {avg_filtered_score:.2f}")
                 else:
                      st.write("Average Overall Score for Filtered Data: N/A (No valid scores found)")


                 # --- Bar Chart: Overall Judgment Distribution (Filtered) ---
                 st.markdown("#### Overall Judgment Distribution (Filtered)")
                 if 'Overall Judgment' in filtered_data.columns and not filtered_data['Overall Judgment'].isna().all():
                      # Ensure non-string types (like None from NaNs) are handled before value_counts
                      valid_judgments = filtered_data['Overall Judgment'].dropna()
                      if not valid_judgments.empty:
                          judgment_counts = valid_judgments.value_counts()
                          # Define a specific order for the judgments
                          judgment_order = [strings["perf_level_very_weak"], strings["perf_level_weak"], strings["perf_level_acceptable"], strings["perf_level_good"], strings["perf_level_excellent"], strings["overall_score_na"]]
                          # Reindex to enforce order, fill missing categories with 0, then drop NA if desired
                          judgment_counts = judgment_counts.reindex(judgment_order, fill_value=0).drop(strings["overall_score_na"], errors='ignore')
                          # Only show the chart if there are counts for actual judgments
                          if not judgment_counts.empty and judgment_counts.sum() > 0:
                               st.bar_chart(judgment_counts)
                          else:
                               st.info("No observations with valid judgments found in the filtered data to chart.")
                      else:
                           st.info("No valid overall judgments found in the filtered data to chart.")
                 else:
                      st.info("Overall Judgment data is not available or is invalid in the filtered dataset.")


                 # --- Bar Chart: Average Score by Domain (Filtered) ---
                 st.markdown("#### Average Score by Domain (Filtered)")
                 # Calculate average for each domain column, ignoring NaNs
                 domain_avg_columns = [col for col in filtered_data.columns if col.startswith('Avg Domain')]
                 if domain_avg_columns:
                      # Select only the domain average columns and calculate the mean for each
                      domain_avg_data = filtered_data[domain_avg_columns].mean().reset_index()
                      domain_avg_data.columns = ['Domain', 'Average Score']
                      # Remove 'Avg ' prefix for cleaner chart labels
                      domain_avg_data['Domain'] = domain_avg_data['Domain'].str.replace('Avg ', '')

                      if not domain_avg_data.empty and not domain_avg_data['Average Score'].isna().all():
                          # Set the index to Domain for st.bar_chart
                          domain_avg_data = domain_avg_data.set_index('Domain')
                          st.bar_chart(domain_avg_data)
                      else:
                          st.info("No valid domain average scores found in the filtered data to chart.")
                 else:
                      st.info("Domain average data is not available in the dataset.")


                 # --- Filtered Data Table and Downloads ---
                 st.markdown("---")
                 st.markdown("##### Filtered Observation Data Table")
                 # Select specific columns for the table display for clarity
                 display_columns = ['Sheet', 'Observer', 'Teacher', 'Operator', 'School', 'Grade', 'Subject', 'Observation Date', 'Overall Score', 'Overall Judgment'] + domain_avg_columns
                 # Ensure selected display columns exist in the filtered data
                 display_columns = [col for col in display_columns if col in filtered_data.columns]
                 st.dataframe(filtered_data[display_columns])

                 st.markdown("###### Download Filtered Data")
                 col_csv, col_excel = st.columns(2)
                 with col_csv:
                      csv_buffer_filtered = io.StringIO()
                      # Include all columns in the downloaded CSV/Excel, not just display columns
                      filtered_data.to_csv(csv_buffer_filtered, index=False)
                      csv_buffer_filtered.seek(0)
                      st.download_button(
                          label=strings["download_filtered_data_csv"],
                          data=csv_buffer_filtered.getvalue(),
                          file_name="filtered_observation_data.csv",
                          mime="text/csv"
                      )
                 with col_excel:
                      excel_buffer_filtered = io.BytesIO()
                      filtered_data.to_excel(excel_buffer_filtered, index=False)
                      excel_buffer_filtered.seek(0)
                      st.download_button(
                           label=strings["download_filtered_data_excel"],
                           data=excel_buffer_filtered.getvalue(),
                           file_name="filtered_observation_data.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                      )

                 st.markdown("---")
                 st.info("""
                     **Note on Chart Downloads:** Streamlit's native charts (like the bar charts above)
                     do not have a built-in image download option. If you require downloadable chart
                     images, you would typically use charting libraries like Matplotlib or Plotly
                     (displayed using `st.pyplot()` or `st.plotly_chart()`) and add specific code
                     for chart export, which can add complexity. The data used to generate the charts
                     is available for download in the filtered data table above.
                     """)


             else:
                 st.info(strings["info_no_observation_data_filtered"]) # Show this if filtered_data is empty


             # --- Overall Domain Averages (Across ALL observations, ignoring filters) ---
             st.markdown("#### Average Score per Domain (Across all observations)")
             # Calculate average for each domain column from the unfiltered data, ignoring NaNs
             domain_avg_columns_all = [col for col in all_obs_data.columns if col.startswith('Avg Domain')]
             if domain_avg_columns_all:
                  domain_avg_data_all = all_obs_data[domain_avg_columns_all].mean().reset_index()
                  domain_avg_data_all.columns = ['Domain', 'Average Score']
                  # Remove 'Avg ' prefix from column names for cleaner chart legend
                  domain_avg_data_all['Domain'] = domain_avg_data_all['Domain'].str.replace('Avg ', '')

                  if not domain_avg_data_all.empty and not domain_avg_data_all['Average Score'].isna().all():
                      # Set the index to Domain for st.bar_chart
                      domain_avg_data_all = domain_avg_data_all.set_index('Domain')
                      st.bar_chart(domain_avg_data_all)

                      st.markdown("###### Download Overall Domain Averages Data")
                      col_csv_all_avg, col_excel_all_avg = st.columns(2)
                      with col_csv_all_avg:
                           csv_buffer_all_avg = io.StringIO()
                           domain_avg_data_all.to_csv(csv_buffer_all_avg)
                           csv_buffer_all_avg.seek(0)
                           st.download_button(
                               label=strings["download_overall_avg_csv"],
                               data=csv_buffer_all_avg.getvalue(),
                               file_name="overall_domain_averages.csv",
                               mime="text/csv"
                           )
                      with col_excel_all_avg:
                           excel_buffer_all_avg = io.BytesIO()
                           domain_avg_data_all.to_excel(excel_buffer_all_avg)
                           excel_buffer_all_avg.seek(0)
                           st.download_button(
                                label=strings["download_overall_avg_excel"],
                                data=excel_buffer_all_avg.getvalue(),
                                file_name="overall_domain_averages.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                           )
                  else:
                       st.info(strings["info_no_numeric_scores_overall"]) # No valid overall domain averages found
             else:
                  st.info("Domain average data columns not found in the dataset.")


             # --- Teacher Performance Over Time ---
             st.markdown("---")
             st.subheader(strings["subheader_teacher_performance"])
             st.info(strings["info_select_teacher"])

             # Teacher selection dropdown for detailed trend
             # Use the list of unique teachers from *all* data, not just filtered data,
             # so you can select any teacher even if they are filtered out by other criteria initially.
             selected_teacher_for_trend = st.selectbox("Select Teacher for Trend Analysis", [None] + all_teachers, format_func=lambda x: x if x is not None else "Select a Teacher...") # Use None for the initial placeholder


             if selected_teacher_for_trend:
                  # Filter the *filtered* data by the selected teacher for the trend analysis
                  teacher_data_for_trend = filtered_data[filtered_data['Teacher'] == selected_teacher_for_trend].copy()

                  if not teacher_data_for_trend.empty:
                       # Display average for the selected teacher within the current filters
                       if 'Overall Score' in teacher_data_for_trend.columns and not teacher_data_for_trend['Overall Score'].isna().all():
                           st.subheader(strings["subheader_teacher_overall_avg"].format(selected_teacher_for_trend))
                           avg_teacher_score_filtered = teacher_data_for_trend['Overall Score'].mean()
                           st.write(f"Average Overall Score (Filtered): {avg_teacher_score_filtered:.2f}")
                       else:
                           st.write(f"Average Overall Score for {selected_teacher_for_trend} (Filtered): N/A (No valid scores found)")


                       # Plot trend over time (Requires valid dates and domain data)
                       domain_avg_columns_teacher = [col for col in teacher_data_for_trend.columns if col.startswith('Avg Domain')]

                       if 'Observation Date' in teacher_data_for_trend.columns and not teacher_data_for_trend['Observation Date'].isna().all() and domain_avg_columns_teacher:
                            st.subheader(strings["subheader_teacher_domain_trend"].format(selected_teacher_for_trend))

                            # Prepare data for plotting trend - requires dates as index and numeric columns
                            # Sort data by date for the trend line
                            trend_data = teacher_data_for_trend.sort_values(by='Observation Date').copy()
                            # Select date and domain average columns
                            trend_columns = ['Observation Date'] + domain_avg_columns_teacher
                            trend_data = trend_data[trend_columns].dropna(subset=['Observation Date']) # Drop rows with no date
                            # Drop domain columns where all values are NaN for this teacher
                            trend_data = trend_data.dropna(axis=1, how='all')


                            if not trend_data.empty and len(trend_data) > 1 and trend_data.select_dtypes(include=np.number).columns.tolist(): # Need at least 2 points and numeric columns to plot a trend
                                 # Set date as index for plotting
                                 trend_data = trend_data.set_index('Observation Date')
                                 # Remove 'Avg ' prefix from column names for cleaner plot legend
                                 trend_data.columns = trend_data.columns.str.replace('Avg ', '')

                                 st.line_chart(trend_data)

                                 st.markdown("###### Download Teacher Trend Data")
                                 col_csv_trend, col_excel_trend = st.columns(2)
                                 with col_csv_trend:
                                      csv_buffer_trend = io.StringIO()
                                      trend_data.to_csv(csv_buffer_trend)
                                      csv_buffer_trend.seek(0)
                                      st.download_button(
                                           label="📥 Download Trend Data (CSV)",
                                           data=csv_buffer_trend.getvalue(),
                                           file_name=f"{selected_teacher_for_trend.replace(' ', '_').lower()}_trend_data.csv",
                                           mime="text/csv"
                                      )
                                 with col_excel_trend:
                                      excel_buffer_trend = io.BytesIO()
                                      trend_data.to_excel(excel_buffer_trend)
                                      excel_buffer_trend.seek(0)
                                      st.download_button(
                                           label="📥 Download Trend Data (Excel)",
                                           data=excel_buffer_trend.getvalue(),
                                           file_name=f"{selected_teacher_for_trend.replace(' ', '_').lower()}_trend_data.xlsx",
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                      )

                            elif len(trend_data) <= 1:
                                st.info("Need at least 2 observations with valid dates and domain scores for this teacher under current filters to show a trend.")
                            elif not trend_data.select_dtypes(include=np.number).columns.tolist():
                                st.info("No numeric domain scores found for this teacher under current filters to show a trend.")
                            else: # Should cover cases where data is present but plotting fails for other reasons
                                st.info("Could not generate trend chart for this teacher with the available data under current filters.")


                       elif 'Observation Date' not in teacher_data_for_trend.columns or teacher_data_for_trend['Observation Date'].isna().all():
                            st.info("Observation dates are missing or invalid for this teacher under current filters.")
                       elif not domain_avg_columns_teacher:
                            st.info("Domain average data is not available for trend analysis.")


                  else:
                       st.info(strings["info_no_obs_for_teacher"])
             else:
                  st.info("Select a teacher above to view their performance trend.")


        else: # If all_obs_data is empty after initial loading
            st.info(strings["info_no_observation_data_filtered"]) # Indicates no LO sheets or failed loading


    # <--- End of Lesson Observation Analytics Page Logic --->


    elif page == strings["page_help"]: # New Help/Guidelines page
        st.title(strings["title_help"])

        st.markdown("---")
        st.markdown("### Using the Application")
        st.markdown("""
        This tool allows you to record lesson observations using a standardized rubric and analyze the collected data.

        **Lesson Observation Input Page:**
        - **Workbook Loading:** The app attempts to load the default Excel workbook (`Teaching Rubric Tool_WeekTemplate.xlsx`). If not found, you'll be prompted. Ensure this file is in the same directory as the script or accessible path.
        - **Sheet Selection:** Choose an existing observation sheet ("LO #") or select "Create new" to start a fresh observation. When creating new, the app finds the next available "LO" number and copies the template sheet.
        - **Cleanup:** The "Clean up unused LO sheets" checkbox allows you to remove sheets that start with "LO " but do not have an Observer Name filled in Cell AA1. This helps keep your workbook organized.
        - **Data Entry:** Fill in the teacher details, lesson information, and rubric scores and notes for each element. Use the rubric guidance expanders to understand the criteria for each element.
        - **Generate Feedback Report:** Check the box if you want a PDF feedback report to be generated upon saving.
        - **Save Observation:** Click "Save Observation" to write the data to the selected (or newly created) sheet in the Excel workbook. The app will perform basic validation. It also updates a 'Feedback Log' sheet with key information and calculates overall/domain scores.
        - **Downloads:** After saving, you can download the updated Excel workbook, the generated PDF feedback report (if selected), and a CSV export of the Feedback Log.

        **Lesson Observation Analytics Page:**
        - This page reads data from all sheets starting with "LO " (excluding the template) and the "Feedback Log" sheet in the loaded workbook.
        - It displays overall statistics and allows you to filter the data by various criteria (School, Grade, Subject, Operator, Teacher, Observer, and Date Range).
        - You can view charts showing overall judgment distribution and average scores per domain for the filtered data, as well as overall domain averages across all observations.
        - A data table of the filtered observations is shown, with options to download it as CSV or Excel.
        - The Teacher Performance Over Time section allows you to select a specific teacher from the filtered data and view their domain score trends over time.

        **App Information and Guidelines Page:**
        - This page provides general information about the app and displays the observation guidelines read directly from the 'Guidelines' sheet in the Excel workbook.

        """)

        st.markdown("---")
        st.markdown("### Observation Guidelines")
        # Read and display guidelines from the Excel sheet
        if wb and "Guidelines" in wb.sheetnames:
            guideline_content = []
            try:
                # Read cells row by row, value only, skip None
                for row in wb["Guidelines"].iter_rows(values_only=True):
                    # Flatten the row and filter out None values, convert to string
                    cleaned_row = [str(cell).strip() for cell in row if cell is not None]
                    guideline_content.extend(cleaned_row) # Add cleaned cells from the row

            except Exception as e:
                st.error(f"Error reading Guidelines sheet: {e}")
                guideline_content = [f"Error loading guidelines: {e}"] # Provide an error message

            # Filter out any empty strings resulting from strip()
            display_content = [line for line in guideline_content if line]

            if display_content:
                # Join content with newlines for Markdown display
                st.markdown("\n".join(display_content))
            else:
                st.info(strings.get("info_no_guidelines", "Guidelines sheet is empty or could not be read."))
        else:
            st.warning("Guidelines sheet not found in the workbook.")


# <--- This 'if wb:' block ends here.
#      The final 'else' block should align with it.
#      This top-level 'if/else' structure handles initial workbook loading errors.
else: # If workbook could not be loaded at the very start
    st.error("Could not load the workbook. Please ensure 'Teaching Rubric Tool_WeekTemplate.xlsx' exists and is accessible.")
